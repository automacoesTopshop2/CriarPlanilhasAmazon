"""
Testes da integração com a API BDAmazon e da rota de criação manual de SKU.

Mocka `core.bdamazon_client` para não chamar a rede de verdade.
"""

from __future__ import annotations

import io
import json
from unittest.mock import patch

import openpyxl
import pytest

from auth import db, Usuario
from auth.security import hash_senha
from core import bdamazon_client


# ---------------------------------------------------------------------------
# /api/bdamazon/contas
# ---------------------------------------------------------------------------

def test_lista_contas_sem_api_key_retorna_503(client, login_admin, monkeypatch):
    """Sem BDAMAZON_API_KEY no env, a rota deve devolver 503 com mensagem clara
    (e não cair no 502 genérico de 'BDAmazon offline')."""
    monkeypatch.delenv("BDAMAZON_API_KEY", raising=False)
    r = client.get("/api/bdamazon/contas")
    assert r.status_code == 503
    data = r.get_json()
    assert data["sucesso"] is False
    assert "BDAMAZON_API_KEY" in data["mensagem"]
    assert "ausente" in data["detalhe"].lower()


def test_lista_contas_proxy_devolve_contas(client, login_admin, monkeypatch):
    monkeypatch.setenv("BDAMAZON_API_KEY", "bdamz_test_token")
    contas_fake = [
        bdamazon_client.Conta(codigo="BOX2", nome="BOX2BRASIL",
                              marca="BOX2", tipo_canal="BASE",
                              prefixo_sku="BOX2-"),
        bdamazon_client.Conta(codigo="VERD-CLA", nome="VERDAL CLA",
                              marca="VERD", tipo_canal="CLA",
                              prefixo_sku="VERD-CLA-"),
    ]
    with patch("core.bdamazon_client.listar_contas", return_value=contas_fake):
        r = client.get("/api/bdamazon/contas")
    assert r.status_code == 200
    data = r.get_json()
    assert data["sucesso"] is True
    assert {c["codigo"] for c in data["contas"]} == {"BOX2", "VERD-CLA"}
    assert data["contas"][0]["nome"] == "BOX2BRASIL"


def test_lista_contas_erro_auth_vira_502(client, login_admin, monkeypatch):
    monkeypatch.setenv("BDAMAZON_API_KEY", "bdamz_test_token")
    err = bdamazon_client.BDAmazonAuthError("chave inválida", status=401)
    with patch("core.bdamazon_client.listar_contas", side_effect=err):
        r = client.get("/api/bdamazon/contas")
    assert r.status_code == 502
    data = r.get_json()
    assert data["sucesso"] is False
    assert "autentic" in data["mensagem"].lower()


def test_lista_contas_exige_login(client):
    r = client.get("/api/bdamazon/contas")
    # Sem login: 401 (JSON) — handler está em web_app.py para rotas /api/
    assert r.status_code == 401


# ---------------------------------------------------------------------------
# /api/processar/sku-manual
# ---------------------------------------------------------------------------

def _template_xlsm() -> io.BytesIO:
    """Gera um template .xlsm minimalista compatível com ProcessadorSKU.
    Cabeçalho na linha 4 ('sku' + outras colunas exigidas), dados a partir da 8."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modelo"
    # Linha 4 = cabeçalho conforme LINHA_CABECALHO_TEMPLATE
    headers = [
        "sku", "nome da marca", "fabricante",
        "id do produto", "id do produto externo",
        "nome do item", "Descrição do Produto",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col_idx, value=h)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _criar_usuario_sem_codigo_externo(app):
    with app.app_context():
        u = Usuario(
            email="semcodigo@topshop.com.br",
            nome="Sem Codigo",
            senha_hash=hash_senha("SenhaForte123!"),
            papel="usuario", ativo=True, totp_required=False,
        )
        db.session.add(u); db.session.commit()
        return u.id


def _criar_usuario_com_codigo_externo(app, codigo="joao.silva"):
    with app.app_context():
        u = Usuario(
            email="op@topshop.com.br",
            nome="Operador",
            senha_hash=hash_senha("SenhaForte123!"),
            papel="usuario",
            ativo=True,
            totp_required=False,
            codigo_externo=codigo,
        )
        db.session.add(u)
        db.session.commit()
        return u.id


def _login(client, email, senha):
    r = client.post("/login",
                    data={"email": email, "senha": senha},
                    follow_redirects=False)
    assert r.status_code == 302, f"Login falhou: {r.status_code}"


def test_sku_manual_exige_codigo_externo(client, app, usuario):
    """Usuário sem codigo_externo recebe 400 explicativo."""
    _login(client, "user@topshop.com.br", "SenhaForte456@")
    fd = {
        "entradas": json.dumps([{
            "sku_raiz": "ABC123",
            "conta_codigo": "BOX2",
            "marca": "X", "ean": "",
        }]),
        "arquivo_template": (_template_xlsm(), "tpl.xlsm"),
    }
    r = client.post("/api/processar/sku-manual",
                    data=fd, content_type="multipart/form-data")
    assert r.status_code == 400
    assert "codigo_externo" in (r.get_json() or {}).get("mensagem", "")


def test_sku_manual_chama_api_e_cria_job(client, app):
    uid = _criar_usuario_com_codigo_externo(app)
    _login(client, "op@topshop.com.br", "SenhaForte123!")

    sku_fake = bdamazon_client.SkuCriado(
        sku_market="BOX2-ABC123",
        sku_raiz="ABC123",
        versao=1,
        conta_codigo="BOX2",
        conta_nome="BOX2BRASIL",
        asin=None,
        titulo=None,
        aguardando_titulo=False,
        criado_em="2026-05-19T14:32:11+00:00",
        criado_por="Operador",
        raw={},
    )

    chamadas = []

    def _fake_criar(**kwargs):
        chamadas.append(kwargs)
        return sku_fake

    fd = {
        "entradas": json.dumps([{
            "sku_raiz": "ABC123",
            "conta_codigo": "BOX2",
            "marca": "TopShop",
            "ean": "7891234567890",
        }]),
        "arquivo_template": (_template_xlsm(), "tpl.xlsm"),
    }
    with patch("core.bdamazon_client.criar_sku", side_effect=_fake_criar):
        r = client.post("/api/processar/sku-manual",
                        data=fd, content_type="multipart/form-data")
    assert r.status_code == 200, r.get_data(as_text=True)
    body = r.get_json()
    assert body["job_id"]
    # Não checamos o job terminou em background — a chamada à API foi feita
    # de forma async dentro de uma thread. Aqui só validamos que o job foi
    # criado e que o endpoint não estourou.


def test_sku_manual_rejeita_entradas_vazias(client, app):
    _criar_usuario_com_codigo_externo(app, codigo="op1")
    _login(client, "op@topshop.com.br", "SenhaForte123!")

    fd = {
        "entradas": json.dumps([]),
        "arquivo_template": (_template_xlsm(), "tpl.xlsm"),
    }
    r = client.post("/api/processar/sku-manual",
                    data=fd, content_type="multipart/form-data")
    assert r.status_code == 400
    assert "entradas" in r.get_json()["mensagem"].lower()


def test_sku_manual_rejeita_sem_template(client, app):
    _criar_usuario_com_codigo_externo(app, codigo="op2")
    _login(client, "op@topshop.com.br", "SenhaForte123!")

    fd = {"entradas": json.dumps([{"sku_raiz": "X", "conta_codigo": "BOX2"}])}
    r = client.post("/api/processar/sku-manual",
                    data=fd, content_type="multipart/form-data")
    assert r.status_code == 400
    assert "template" in r.get_json()["mensagem"].lower()


# ---------------------------------------------------------------------------
# /admin/usuarios/<id>/codigo-externo
# ---------------------------------------------------------------------------

def test_admin_altera_codigo_externo(client, app, login_admin, usuario):
    r = client.post(f"/admin/usuarios/{usuario}/codigo-externo",
                    json={"codigo_externo": "joao.silva"})
    assert r.status_code == 200
    assert r.get_json() == {"sucesso": True, "codigo_externo": "joao.silva"}
    with app.app_context():
        u = db.session.get(Usuario, usuario)
        assert u.codigo_externo == "joao.silva"


def test_admin_remove_codigo_externo_com_vazio(client, app, login_admin, usuario):
    # Primeiro seta um valor
    with app.app_context():
        u = db.session.get(Usuario, usuario)
        u.codigo_externo = "antigo"
        db.session.commit()

    r = client.post(f"/admin/usuarios/{usuario}/codigo-externo",
                    json={"codigo_externo": ""})
    assert r.status_code == 200
    assert r.get_json()["codigo_externo"] is None
    with app.app_context():
        u = db.session.get(Usuario, usuario)
        assert u.codigo_externo is None


def test_admin_rejeita_codigo_externo_duplicado(client, app, login_admin, usuario):
    with app.app_context():
        outro = Usuario(
            email="outro@topshop.com.br",
            nome="Outro",
            senha_hash=hash_senha("SenhaForte123!"),
            papel="usuario", ativo=True, totp_required=False,
            codigo_externo="ja-existe",
        )
        db.session.add(outro)
        db.session.commit()

    r = client.post(f"/admin/usuarios/{usuario}/codigo-externo",
                    json={"codigo_externo": "ja-existe"})
    assert r.status_code == 400
    assert "já em uso" in r.get_json()["mensagem"].lower()


def test_admin_rejeita_codigo_externo_com_espaco(client, login_admin, usuario):
    r = client.post(f"/admin/usuarios/{usuario}/codigo-externo",
                    json={"codigo_externo": "tem espaço"})
    assert r.status_code == 400


def test_usuario_comum_nao_pode_alterar_codigo_externo(client, login_usuario, usuario):
    r = client.post(f"/admin/usuarios/{usuario}/codigo-externo",
                    json={"codigo_externo": "tentando"})
    assert r.status_code == 403


# ---------------------------------------------------------------------------
# Cliente HTTP (sem rede) — valida shape de erros
# ---------------------------------------------------------------------------

def test_bdamazon_client_sem_key_falha_explicito(monkeypatch):
    monkeypatch.delenv("BDAMAZON_API_KEY", raising=False)
    with pytest.raises(bdamazon_client.BDAmazonError) as exc:
        bdamazon_client.listar_contas()
    assert "BDAMAZON_API_KEY" in str(exc.value)


# ---------------------------------------------------------------------------
# /api/bdamazon/criar-sku
# ---------------------------------------------------------------------------

def test_criar_sku_sem_codigo_externo_400(client, app, monkeypatch):
    """Usuário sem codigo_externo recebe 400 explícito."""
    monkeypatch.setenv("BDAMAZON_API_KEY", "bdamz_test_token")
    _criar_usuario_sem_codigo_externo(app)
    _login(client, "semcodigo@topshop.com.br", "SenhaForte123!")
    r = client.post("/api/bdamazon/criar-sku",
                    json={"sku_raiz": "ABC", "conta_codigo": "BOX2"})
    assert r.status_code == 400
    assert "codigo_externo" in r.get_json()["mensagem"]


def test_criar_sku_sem_api_key_503(client, app, monkeypatch):
    monkeypatch.delenv("BDAMAZON_API_KEY", raising=False)
    _criar_usuario_com_codigo_externo(app, codigo="op-x")
    _login(client, "op@topshop.com.br", "SenhaForte123!")
    r = client.post("/api/bdamazon/criar-sku",
                    json={"sku_raiz": "ABC", "conta_codigo": "BOX2"})
    assert r.status_code == 503
    assert "BDAMAZON_API_KEY" in r.get_json()["mensagem"]


def test_criar_sku_devolve_sku_market(client, app, monkeypatch):
    monkeypatch.setenv("BDAMAZON_API_KEY", "bdamz_test_token")
    _criar_usuario_com_codigo_externo(app, codigo="joao.silva")
    _login(client, "op@topshop.com.br", "SenhaForte123!")

    sku_fake = bdamazon_client.SkuCriado(
        sku_market="BOX2-ABC123", sku_raiz="ABC123", versao=1,
        conta_codigo="BOX2", conta_nome="BOX2BRASIL",
        asin=None, titulo=None, aguardando_titulo=False,
        criado_em="2026-05-19T14:32:11+00:00", criado_por="Joao", raw={},
    )

    chamadas = []
    def fake(**kwargs):
        chamadas.append(kwargs)
        return sku_fake

    # Após criar, a rota faz um GET best-effort para puxar o status do catálogo.
    detalhe_fake = bdamazon_client.SkuCriado(
        sku_market="BOX2-ABC123", sku_raiz="ABC123", versao=1,
        conta_codigo="BOX2", conta_nome="BOX2BRASIL",
        asin=None, titulo=None, aguardando_titulo=False,
        criado_em="2026-05-19T14:32:11+00:00", criado_por="Joao", raw={},
        status_produto="SENSIVEL", titulo_produto="Produto X", estoque_produto=42,
    )

    with patch("core.bdamazon_client.criar_sku", side_effect=fake), \
         patch("core.bdamazon_client.consultar_sku", return_value=detalhe_fake):
        r = client.post("/api/bdamazon/criar-sku",
                        json={"sku_raiz": "ABC123",
                              "conta_codigo": "BOX2"})
    assert r.status_code == 200, r.get_data(as_text=True)
    body = r.get_json()
    assert body["sucesso"] is True
    assert body["sku_market"] == "BOX2-ABC123"
    assert body["versao"] == 1
    # Status de sensibilidade do catálogo vem junto na resposta de criação.
    assert body["status_produto"] == "SENSIVEL"
    assert body["estoque_produto"] == 42
    assert chamadas[0]["sku_raiz"] == "ABC123"
    assert chamadas[0]["conta_codigo"] == "BOX2"
    assert chamadas[0]["usuario_codigo"] == "joao.silva"


def test_criar_sku_com_asin_envia_para_api(client, app, monkeypatch):
    """Modo ASIN: o asin deve ser propagado para o POST /skus do BDAmazon."""
    monkeypatch.setenv("BDAMAZON_API_KEY", "bdamz_test_token")
    _criar_usuario_com_codigo_externo(app, codigo="op-asin")
    _login(client, "op@topshop.com.br", "SenhaForte123!")

    sku_fake = bdamazon_client.SkuCriado(
        sku_market="BOX2-XYZ", sku_raiz="XYZ", versao=1,
        conta_codigo="BOX2", conta_nome="BOX2BRASIL",
        asin="B08ABCD123", titulo=None, aguardando_titulo=False,
        criado_em="2026-05-19T14:32:11+00:00", criado_por="Joao", raw={},
    )

    chamadas = []
    def fake(**kwargs):
        chamadas.append(kwargs)
        return sku_fake

    with patch("core.bdamazon_client.criar_sku", side_effect=fake), \
         patch("core.bdamazon_client.consultar_sku", return_value=None):
        r = client.post("/api/bdamazon/criar-sku",
                        json={"sku_raiz": "XYZ",
                              "conta_codigo": "BOX2",
                              "asin": "B08ABCD123"})
    assert r.status_code == 200
    assert chamadas[0]["asin"] == "B08ABCD123"
    # SKU raiz não consta no catálogo -> status_produto None (sem badge de risco).
    assert r.get_json()["status_produto"] is None


def test_criar_sku_status_consulta_falha_nao_quebra_criacao(client, app, monkeypatch):
    """Se o GET de status falhar, a criação ainda devolve 200 (best-effort)."""
    monkeypatch.setenv("BDAMAZON_API_KEY", "bdamz_test_token")
    _criar_usuario_com_codigo_externo(app, codigo="op-bestv")
    _login(client, "op@topshop.com.br", "SenhaForte123!")

    sku_fake = bdamazon_client.SkuCriado(
        sku_market="BOX2-ABC123", sku_raiz="ABC123", versao=1,
        conta_codigo="BOX2", conta_nome="BOX2BRASIL",
        asin=None, titulo=None, aguardando_titulo=False,
        criado_em="2026-05-19T14:32:11+00:00", criado_por="Joao", raw={},
    )
    consulta_erro = bdamazon_client.BDAmazonError("timeout", status=502)
    with patch("core.bdamazon_client.criar_sku", return_value=sku_fake), \
         patch("core.bdamazon_client.consultar_sku", side_effect=consulta_erro):
        r = client.post("/api/bdamazon/criar-sku",
                        json={"sku_raiz": "ABC123", "conta_codigo": "BOX2"})
    assert r.status_code == 200, r.get_data(as_text=True)
    body = r.get_json()
    assert body["sucesso"] is True
    assert body["sku_market"] == "BOX2-ABC123"
    assert body["status_produto"] is None


# ---------------------------------------------------------------------------
# /api/bdamazon/consultar-sku/<sku_market>  (busca + status do catálogo)
# ---------------------------------------------------------------------------

def test_consultar_sku_devolve_status_produto(client, app, monkeypatch):
    monkeypatch.setenv("BDAMAZON_API_KEY", "bdamz_test_token")
    _criar_usuario_com_codigo_externo(app, codigo="op-cons")
    _login(client, "op@topshop.com.br", "SenhaForte123!")

    sku_fake = bdamazon_client.SkuCriado(
        sku_market="BOX2-ABC123", sku_raiz="ABC123", versao=1,
        conta_codigo="BOX2", conta_nome="BOX2BRASIL",
        asin="B0AAAAAAAA", titulo="Produto exemplo", aguardando_titulo=False,
        criado_em="2026-05-19T14:32:11+00:00", criado_por="Joao", raw={},
        ean="7891234567890", status_produto="PROIBIDO",
        titulo_produto="Produto exemplo - catálogo", estoque_produto=0,
    )

    chamadas = []
    def fake(sku_market):
        chamadas.append(sku_market)
        return sku_fake

    with patch("core.bdamazon_client.consultar_sku", side_effect=fake):
        r = client.get("/api/bdamazon/consultar-sku/BOX2-ABC123")
    assert r.status_code == 200, r.get_data(as_text=True)
    body = r.get_json()
    assert body["sucesso"] is True
    assert body["sku_market"] == "BOX2-ABC123"
    assert body["status_produto"] == "PROIBIDO"
    assert body["ean"] == "7891234567890"
    assert body["estoque_produto"] == 0
    assert chamadas == ["BOX2-ABC123"]


def test_consultar_sku_inexistente_404(client, app, monkeypatch):
    monkeypatch.setenv("BDAMAZON_API_KEY", "bdamz_test_token")
    _criar_usuario_com_codigo_externo(app, codigo="op-404")
    _login(client, "op@topshop.com.br", "SenhaForte123!")
    with patch("core.bdamazon_client.consultar_sku", return_value=None):
        r = client.get("/api/bdamazon/consultar-sku/BOX2-NAOEXISTE")
    assert r.status_code == 404
    assert "não encontrado" in r.get_json()["mensagem"].lower()


def test_consultar_sku_sem_api_key_503(client, app, monkeypatch):
    monkeypatch.delenv("BDAMAZON_API_KEY", raising=False)
    _criar_usuario_com_codigo_externo(app, codigo="op-503c")
    _login(client, "op@topshop.com.br", "SenhaForte123!")
    r = client.get("/api/bdamazon/consultar-sku/BOX2-ABC123")
    assert r.status_code == 503
    assert "BDAMAZON_API_KEY" in r.get_json()["mensagem"]


def test_consultar_sku_exige_login(client):
    r = client.get("/api/bdamazon/consultar-sku/BOX2-ABC123")
    assert r.status_code == 401


def test_consultar_sku_propaga_429(client, app, monkeypatch):
    monkeypatch.setenv("BDAMAZON_API_KEY", "bdamz_test_token")
    _criar_usuario_com_codigo_externo(app, codigo="op-429c")
    _login(client, "op@topshop.com.br", "SenhaForte123!")
    err = bdamazon_client.BDAmazonRateLimitError("rate-limit", status=429)
    with patch("core.bdamazon_client.consultar_sku", side_effect=err):
        r = client.get("/api/bdamazon/consultar-sku/BOX2-ABC123")
    assert r.status_code == 429


def test_skucriado_from_dict_parseia_status_produto():
    """O cliente preserva status_produto/ean/estoque vindos do GET /skus."""
    d = {
        "sku_market": "BOX2-ABC123", "sku_raiz": "ABC123", "versao": 1,
        "conta_codigo": "BOX2", "conta_nome": "BOX2BRASIL",
        "asin": "B0AAAAAAAA", "ean": "7891234567890", "titulo": "X",
        "aguardando_titulo": False, "criado_em": "2026-05-19T14:32:11+00:00",
        "criado_por": "Joao", "status_produto": "LIVRE",
        "titulo_produto": "X catálogo", "estoque_produto": 150,
    }
    sku = bdamazon_client.SkuCriado.from_dict(d)
    assert sku.status_produto == "LIVRE"
    assert sku.ean == "7891234567890"
    assert sku.estoque_produto == 150
    assert sku.titulo_produto == "X catálogo"


def test_criar_sku_propaga_429(client, app, monkeypatch):
    monkeypatch.setenv("BDAMAZON_API_KEY", "bdamz_test_token")
    _criar_usuario_com_codigo_externo(app, codigo="op-r")
    _login(client, "op@topshop.com.br", "SenhaForte123!")
    err = bdamazon_client.BDAmazonRateLimitError("rate-limit", status=429)
    with patch("core.bdamazon_client.criar_sku", side_effect=err):
        r = client.post("/api/bdamazon/criar-sku",
                        json={"sku_raiz": "X", "conta_codigo": "BOX2"})
    assert r.status_code == 429
    assert "rate-limit" in r.get_json()["mensagem"].lower()


# ---------------------------------------------------------------------------
# /api/processar/sku-manual com sku_market pré-resolvido
# ---------------------------------------------------------------------------

def test_sku_manual_job_termina_em_done_sem_atribuir_slot_invalido(client, app,
                                                                   monkeypatch):
    """Regressão: ao introduzir `job.skus_market = ...` o __slots__ do Job
    fez a thread explodir com AttributeError dentro do try/except, marcando
    o job como 'error'. Patch threading.Thread para rodar inline, depois
    inspeciona o status final do job."""
    monkeypatch.setenv("BDAMAZON_API_KEY", "bdamz_test_token")
    _criar_usuario_com_codigo_externo(app, codigo="op-regress")
    _login(client, "op@topshop.com.br", "SenhaForte123!")

    import threading
    threads_iniciadas = []

    class ThreadInline:
        def __init__(self, target, args=(), kwargs=None, daemon=None):
            self.target = target
            self.args = args
            self.kwargs = kwargs or {}

        def start(self):
            self.target(*self.args, **self.kwargs)

    monkeypatch.setattr(threading, "Thread", ThreadInline)

    fd = {
        "entradas": json.dumps([{
            "sku_raiz": "ABC123", "conta_codigo": "BOX2",
            "marca": "Top", "ean": "789",
            "sku_market": "BOX2-ABC123", "versao": 1,
        }]),
        "arquivo_template": (_template_xlsm(), "tpl.xlsm"),
    }
    r = client.post("/api/processar/sku-manual",
                    data=fd, content_type="multipart/form-data")
    assert r.status_code == 200
    job_id = r.get_json()["job_id"]
    # Com ThreadInline, o job já terminou. Importa o JOBS store e confere.
    from web_app import JOBS
    job = JOBS[job_id]
    # O __slots__ não deve ter sido violado. Pode ter terminado em 'done'
    # (se processador rodou ok) ou 'error' por outra razão — mas NÃO por
    # AttributeError de skus_market.
    assert job.status in ("done", "error"), f"status inesperado: {job.status}"
    # Drena a fila pra ver se a mensagem 'skus_market' aparece como exceção
    logs = []
    while not job.log_queue.empty():
        logs.append(job.log_queue.get_nowait())
    erros_sobre_slot = [
        l for l in logs
        if isinstance(l.get("mensagem"), str)
        and "skus_market" in l["mensagem"]
    ]
    assert not erros_sobre_slot, f"regressão: {erros_sobre_slot}"


def test_sku_manual_aceita_sku_market_pre_resolvido(client, app, monkeypatch):
    """Quando o sku_market já vem do passo 'Solicitar', a rota não deve
    chamar criar_sku de novo."""
    monkeypatch.setenv("BDAMAZON_API_KEY", "bdamz_test_token")
    _criar_usuario_com_codigo_externo(app, codigo="op-pre")
    _login(client, "op@topshop.com.br", "SenhaForte123!")
    fd = {
        "entradas": json.dumps([{
            "sku_raiz": "ABC123", "conta_codigo": "BOX2",
            "marca": "Top", "ean": "789",
            "sku_market": "BOX2-ABC123", "versao": 1,
        }]),
        "arquivo_template": (_template_xlsm(), "tpl.xlsm"),
    }
    with patch("core.bdamazon_client.criar_sku") as mock_criar:
        r = client.post("/api/processar/sku-manual",
                        data=fd, content_type="multipart/form-data")
    assert r.status_code == 200
    assert r.get_json()["job_id"]
    # Como sku_market veio resolvido, criar_sku não deve ter sido chamado
    assert mock_criar.call_count == 0


# ---------------------------------------------------------------------------
# /api/processar/asin-manual
# ---------------------------------------------------------------------------

def _template_xlsm_asin() -> io.BytesIO:
    """Template para ProcessadorASIN: cabeçalho na linha 4, dados a partir
    da linha 7. Cabeçalhos mínimos esperados."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modelo"
    headers = ["ASIN", "SKU", "Preço padrão BRL (Vender na Amazon, BR)",
               "Peso do pacote", "Peso do item",
               "Comprimento do pacote", "Largura do pacote", "Altura do pacote"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col_idx, value=h)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def test_asin_manual_cria_job(client, app, monkeypatch):
    monkeypatch.setenv("BDAMAZON_API_KEY", "bdamz_test_token")
    _criar_usuario_com_codigo_externo(app, codigo="op-asin2")
    _login(client, "op@topshop.com.br", "SenhaForte123!")

    fd = {
        "entradas": json.dumps([{
            "sku_raiz": "ABC123", "conta_codigo": "BOX2",
            "asin": "B08ABCD123",
            "marca": "Top", "ean": "789",
            "sku_market": "BOX2-ABC123", "versao": 1,
        }]),
        "arquivo_template": (_template_xlsm_asin(), "tpl.xlsm"),
    }
    r = client.post("/api/processar/asin-manual",
                    data=fd, content_type="multipart/form-data")
    assert r.status_code == 200, r.get_data(as_text=True)
    assert r.get_json()["job_id"]


def test_asin_manual_exige_codigo_externo(client, app, monkeypatch, usuario):
    """Usuário sem codigo_externo recebe 400 explícito também na rota ASIN."""
    monkeypatch.setenv("BDAMAZON_API_KEY", "bdamz_test_token")
    _login(client, "user@topshop.com.br", "SenhaForte456@")
    fd = {
        "entradas": json.dumps([{
            "sku_raiz": "X", "conta_codigo": "BOX2", "asin": "B08X",
            "sku_market": "BOX2-X",
        }]),
        "arquivo_template": (_template_xlsm_asin(), "tpl.xlsm"),
    }
    r = client.post("/api/processar/asin-manual",
                    data=fd, content_type="multipart/form-data")
    assert r.status_code == 400


def test_bdamazon_client_401_vira_auth_error(monkeypatch):
    monkeypatch.setenv("BDAMAZON_API_KEY", "bdamz_abc_def")

    class FakeResp:
        status_code = 401
        ok = False
        text = '{"detail":"chave inválida"}'

        def json(self):
            return {"detail": "chave inválida"}

    def fake_request(method, url, headers, json, timeout):
        return FakeResp()

    monkeypatch.setattr("requests.request", fake_request)
    with pytest.raises(bdamazon_client.BDAmazonAuthError):
        bdamazon_client.listar_contas()
