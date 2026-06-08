"""
Testes do "template pronto" do Criar por SKU.

Cobre:
    - status em /api/bases (template_sku)
    - upload / delete do template pronto
    - reuso do template salvo via flag `usar_template_salvo` nas rotas de
      processamento (modo planilha e modo manual)
    - persistência do upload quando `salvar_template=1`

DATA_DIR é isolado por teste (tmp_path) para não tocar no volume real.
"""

from __future__ import annotations

import io
import json

import openpyxl
import pytest

from auth import db, Usuario
from auth.security import hash_senha


@pytest.fixture(autouse=True)
def _isolar_data_dir(tmp_path, monkeypatch):
    """Aponta DATA_DIR para um tmp por teste; o template pronto vive lá."""
    monkeypatch.setenv("DATA_DIR", str(tmp_path))
    yield


def _template_xlsm() -> io.BytesIO:
    """Template .xlsm minimalista compatível com ProcessadorSKU."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modelo"
    headers = [
        "sku", "nome da marca", "fabricante",
        "id do produto", "id do produto externo",
        "nome do item", "Descrição do Produto",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col_idx, value=h)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _entrada_sku_xlsx() -> io.BytesIO:
    """Planilha de entrada mínima (uma coluna de SKU)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="SKU")
    ws.cell(row=2, column=1, value="ABC123")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _login(client, email, senha):
    r = client.post("/login", data={"email": email, "senha": senha},
                    follow_redirects=False)
    assert r.status_code == 302, f"Login falhou: {r.status_code}"


def _criar_usuario_com_codigo_externo(app, codigo="op.tpl"):
    with app.app_context():
        u = Usuario(
            email="op@topshop.com.br", nome="Operador",
            senha_hash=hash_senha("SenhaForte123!"),
            papel="usuario", ativo=True, totp_required=False,
            codigo_externo=codigo,
        )
        db.session.add(u)
        db.session.commit()
        return u.id


def _upload_template(client, nome="NOGORA_modelo.xlsm"):
    return client.post(
        "/api/template/sku/upload",
        data={"arquivo": (_template_xlsm(), nome)},
        content_type="multipart/form-data",
    )


# ---------------------------------------------------------------------------
# status / upload / delete
# ---------------------------------------------------------------------------

def test_bases_sem_template_pronto(client, login_usuario):
    r = client.get("/api/bases")
    assert r.status_code == 200
    ts = r.get_json()["template_sku"]
    assert ts["existe"] is False
    assert ts["atualizado_em"] is None


def test_upload_template_salva_e_aparece_no_status(client, login_usuario):
    r = _upload_template(client, "NOGORA_modelo.xlsm")
    assert r.status_code == 200, r.get_data(as_text=True)
    body = r.get_json()
    assert body["sucesso"] is True
    assert body["bases"]["template_sku"]["existe"] is True
    assert body["bases"]["template_sku"]["arquivo"] == "NOGORA_modelo.xlsm"

    # E persiste no GET subsequente
    ts = client.get("/api/bases").get_json()["template_sku"]
    assert ts["existe"] is True
    assert ts["atualizado_em"] is not None
    assert ts["tamanho"] and ts["tamanho"] > 0


def test_upload_rejeita_extensao_invalida(client, login_usuario):
    r = client.post(
        "/api/template/sku/upload",
        data={"arquivo": (io.BytesIO(b"nope"), "planilha.xlsx")},
        content_type="multipart/form-data",
    )
    assert r.status_code == 400
    assert ".xlsm" in r.get_json()["mensagem"]


def test_upload_sem_arquivo_400(client, login_usuario):
    r = client.post("/api/template/sku/upload", data={},
                    content_type="multipart/form-data")
    assert r.status_code == 400


def test_delete_remove_template(client, login_usuario):
    _upload_template(client)
    assert client.get("/api/bases").get_json()["template_sku"]["existe"] is True

    r = client.delete("/api/template/sku")
    assert r.status_code == 200
    assert r.get_json()["bases"]["template_sku"]["existe"] is False
    assert client.get("/api/bases").get_json()["template_sku"]["existe"] is False


def test_template_endpoints_exigem_login(client):
    assert client.post("/api/template/sku/upload").status_code == 401
    assert client.delete("/api/template/sku").status_code == 401


# ---------------------------------------------------------------------------
# reuso na rota de planilha (/api/processar/sku)
# ---------------------------------------------------------------------------

def test_processar_sku_usa_template_salvo(client, login_usuario):
    _upload_template(client)
    fd = {
        "arquivo_entrada": (_entrada_sku_xlsx(), "skus.xlsx"),
        "usar_template_salvo": "1",
    }
    r = client.post("/api/processar/sku", data=fd,
                    content_type="multipart/form-data")
    assert r.status_code == 200, r.get_data(as_text=True)
    assert r.get_json()["job_id"]


def test_processar_sku_sem_template_salvo_400(client, login_usuario):
    fd = {
        "arquivo_entrada": (_entrada_sku_xlsx(), "skus.xlsx"),
        "usar_template_salvo": "1",
    }
    r = client.post("/api/processar/sku", data=fd,
                    content_type="multipart/form-data")
    assert r.status_code == 400
    assert "template pronto" in r.get_data(as_text=True).lower()


def test_processar_sku_salva_template_quando_pedido(client, login_usuario):
    fd = {
        "arquivo_entrada": (_entrada_sku_xlsx(), "skus.xlsx"),
        "arquivo_template": (_template_xlsm(), "NOGORA_novo.xlsm"),
        "salvar_template": "1",
    }
    r = client.post("/api/processar/sku", data=fd,
                    content_type="multipart/form-data")
    assert r.status_code == 200
    # O upload virou o template pronto
    ts = client.get("/api/bases").get_json()["template_sku"]
    assert ts["existe"] is True
    assert ts["arquivo"] == "NOGORA_novo.xlsm"


# ---------------------------------------------------------------------------
# reuso na rota manual (/api/processar/sku-manual)
# ---------------------------------------------------------------------------

def test_sku_manual_usa_template_salvo(client, app):
    _criar_usuario_com_codigo_externo(app, codigo="op-tpl-m")
    _login(client, "op@topshop.com.br", "SenhaForte123!")
    _upload_template(client)

    fd = {
        "entradas": json.dumps([{
            "sku_raiz": "ABC123", "conta_codigo": "BOX2",
            "marca": "Top", "ean": "789",
            "sku_market": "BOX2-ABC123", "versao": 1,
        }]),
        "usar_template_salvo": "1",
    }
    r = client.post("/api/processar/sku-manual", data=fd,
                    content_type="multipart/form-data")
    assert r.status_code == 200, r.get_data(as_text=True)
    assert r.get_json()["job_id"]


def test_sku_manual_usar_salvo_sem_template_400(client, app):
    _criar_usuario_com_codigo_externo(app, codigo="op-tpl-m2")
    _login(client, "op@topshop.com.br", "SenhaForte123!")

    fd = {
        "entradas": json.dumps([{
            "sku_raiz": "ABC123", "conta_codigo": "BOX2",
            "sku_market": "BOX2-ABC123", "versao": 1,
        }]),
        "usar_template_salvo": "1",
    }
    r = client.post("/api/processar/sku-manual", data=fd,
                    content_type="multipart/form-data")
    assert r.status_code == 400
    assert "template pronto" in r.get_json()["mensagem"].lower()
