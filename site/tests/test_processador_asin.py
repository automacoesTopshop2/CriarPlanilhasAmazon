"""
Testes do ProcessadorASIN.

Foco: garantir que campos custom adicionados ao valores_fixos_padrao
(via /admin/configuracoes) sejam aplicados na planilha gerada — mesmo
bug que fez o usuário pensar que só /sku consumia as customizações.
"""

from __future__ import annotations

import io

import openpyxl
import pytest

from core.config import Configuracoes
from core.processadores.asin import ProcessadorASIN


@pytest.fixture(autouse=True)
def _forcar_descricao_planilha(monkeypatch):
    """Estes testes validam a leitura da planilha de Descrição local. Como o
    padrão do sistema passou a ser a API do AgentedeTitulos, força aqui o
    modo-planilha determinístico (sem key), independente do ambiente."""
    monkeypatch.delenv("TITULOS_API_KEY", raising=False)
    monkeypatch.setenv("USAR_PLANILHA_DESCRICAO", "1")


def _template_asin_com_colunas(cabecalhos):
    """Gera template ListaASINS in-memory com os cabeçalhos pedidos na linha 4."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modelo"
    for col_idx, h in enumerate(cabecalhos, start=1):
        ws.cell(row=4, column=col_idx, value=h)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def _entrada_asin(linhas):
    """Gera planilha de entrada ASIN/SKU."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ASIN", "SKU"])
    for asin, sku in linhas:
        ws.append([asin, sku])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def test_asin_escreve_campos_fixos_padrao(tmp_path):
    """Smoke: valores_fixos_padrao default são escritos no template."""
    cfg = Configuracoes()
    # Bases vazias — não vamos validar preço/descricao aqui
    cfg.arquivo_precificacao = str(tmp_path / "p.xlsx")
    cfg.arquivo_descricao = str(tmp_path / "d.xlsx")
    _wb_vazio_precificacao(cfg.arquivo_precificacao)
    _wb_vazio_descricao(cfg.arquivo_descricao)

    proc = ProcessadorASIN(cfg)
    template = _template_asin_com_colunas([
        "ASIN", "SKU", "Condição do Produto", "País de origem",
    ])
    entrada = _entrada_asin([("B08X", "BOX2-ABC")])
    resultado = proc.processar(arquivo_entrada=entrada, arquivo_template=template)
    assert resultado.sucesso, resultado.mensagem

    wb = openpyxl.load_workbook(resultado.arquivo_saida)
    ws = wb.active
    # Linha 7 = LINHA_INICIO_DADOS_PADRAO do ProcessadorASIN
    assert ws.cell(row=7, column=1).value == "B08X"
    assert ws.cell(row=7, column=2).value == "BOX2-ABC"
    assert ws.cell(row=7, column=3).value == "Novo"   # Condição do Produto
    assert ws.cell(row=7, column=4).value == "Brasil" # País de origem


def test_asin_aplica_campo_custom_da_config(tmp_path):
    """REGRESSÃO: campo novo em valores_fixos_padrao tem que ir pra planilha ASIN.

    Antes, o ProcessadorASIN iterava uma lista hardcoded de 12 campos —
    qualquer chave nova adicionada via /admin/configuracoes ficava de fora,
    fazendo o admin pensar que 'só o SKU pega as mudanças'."""
    cfg = Configuracoes()
    cfg.arquivo_precificacao = str(tmp_path / "p.xlsx")
    cfg.arquivo_descricao = str(tmp_path / "d.xlsx")
    _wb_vazio_precificacao(cfg.arquivo_precificacao)
    _wb_vazio_descricao(cfg.arquivo_descricao)

    # Simula o que aplicar_gerenciador faz quando o admin cadastra um campo
    # com nome novo (ex.: a Amazon renomeou uma coluna no template ListaASINS).
    cfg.valores_fixos_padrao["Categoria fiscal personalizada (BR)"] = "9504.50.90"
    cfg.valores_fixos_padrao["Origem da mercadoria (BR)"] = "0"

    proc = ProcessadorASIN(cfg)
    template = _template_asin_com_colunas([
        "ASIN", "SKU",
        "Categoria fiscal personalizada (BR)",
        "Origem da mercadoria (BR)",
    ])
    entrada = _entrada_asin([("B09Y", "NOGO-XYZ")])
    resultado = proc.processar(arquivo_entrada=entrada, arquivo_template=template)
    assert resultado.sucesso, resultado.mensagem

    wb = openpyxl.load_workbook(resultado.arquivo_saida)
    ws = wb.active
    assert ws.cell(row=7, column=3).value == "9504.50.90"
    assert ws.cell(row=7, column=4).value == "0"


def test_asin_respeita_campo_custom_sobrescrevendo_default(tmp_path):
    """Custom com mesma chave que um default deve vencer (aplicar_gerenciador
    faz update sobrescrevendo)."""
    cfg = Configuracoes()
    cfg.arquivo_precificacao = str(tmp_path / "p.xlsx")
    cfg.arquivo_descricao = str(tmp_path / "d.xlsx")
    _wb_vazio_precificacao(cfg.arquivo_precificacao)
    _wb_vazio_descricao(cfg.arquivo_descricao)

    cfg.valores_fixos_padrao["País de origem"] = "China"  # override

    proc = ProcessadorASIN(cfg)
    template = _template_asin_com_colunas(["ASIN", "SKU", "País de origem"])
    entrada = _entrada_asin([("B0", "X")])
    resultado = proc.processar(arquivo_entrada=entrada, arquivo_template=template)
    assert resultado.sucesso
    wb = openpyxl.load_workbook(resultado.arquivo_saida)
    assert wb.active.cell(row=7, column=3).value == "China"


# ---------------------------------------------------------------------------
# Helpers: planilhas-base mínimas para o ProcessadorASIN carregar sem erro
# ---------------------------------------------------------------------------

def _wb_vazio_precificacao(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["SKU", "Padrão"])  # cabeçalhos mínimos
    wb.save(path)


def _wb_vazio_descricao(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["SKU"])
    wb.save(path)
