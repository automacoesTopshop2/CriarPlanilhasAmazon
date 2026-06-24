# ==============================================================================
# CARREGADOR DE DESCRIÇÃO
# ==============================================================================
# Classe responsável por carregar e gerenciar a base de dados de descrições.
#
# Funcionalidades:
#   - Carrega planilha de descrições de produtos
#   - Mapeia dinamicamente colunas por nome
#   - Suporta cálculo de Kits (K2-, K3-, K-)
#   - Fornece dados como título, peso, medidas, tópicos, etc.
# ==============================================================================

import os
import re
import openpyxl
from typing import Dict, Optional, Any, List, Tuple
from dataclasses import dataclass, field
from datetime import datetime

from ..utils import Utilitarios
from ..config import Configuracoes


@dataclass
class DadosProduto:
    """
    Estrutura para armazenar os dados de descrição de um produto.
    
    Attributes:
        sku_base: SKU base do produto
        titulo: Título/Nome do produto
        descricao: Descrição completa
        modelo: Número do modelo/referência
        ean: Código EAN/GTIN
        peso: Peso em quilogramas
        comprimento: Comprimento em centímetros
        largura: Largura em centímetros
        altura: Altura em centímetros
        topicos: Lista de tópicos/bullets
    """
    sku_base: str
    titulo: str = ""
    descricao: str = ""
    modelo: str = ""
    ean: str = ""
    peso: str = ""
    comprimento: str = ""
    largura: str = ""
    altura: str = ""
    topicos: List[str] = field(default_factory=list)
    
    def para_dicionario(self) -> Dict[str, Any]:
        """Converte para dicionário compatível com o sistema legado."""
        return {
            'P': self.titulo,
            'desc': self.descricao,
            'modelo': self.modelo,
            'ean': self.ean,
            'peso': self.peso,
            'comp': self.comprimento,
            'larg': self.largura,
            'alt': self.altura,
            'topico1': self.topicos[0] if len(self.topicos) > 0 else "",
            'topico2': self.topicos[1] if len(self.topicos) > 1 else "",
            'topico3': self.topicos[2] if len(self.topicos) > 2 else "",
            'topico4': self.topicos[3] if len(self.topicos) > 3 else "",
            'topico5': self.topicos[4] if len(self.topicos) > 4 else "",
        }


class CarregadorDescricao:
    """
    Carrega e gerencia a base de dados de descrições de produtos.
    
    Esta classe é responsável por:
    - Carregar a planilha de descrições
    - Mapear dinamicamente colunas por nome
    - Calcular dados de Kits (K2-, K3-, K-)
    - Fornecer acesso aos dados de cada produto
    
    Attributes:
        config: Instância de Configuracoes com os parâmetros
        dados: Dicionário com os dados carregados
        _carregado: Flag indicando se os dados foram carregados
    """
    
    # Nome da aba preferida na planilha de descrição
    ABA_PREFERIDA = "PRODUTO SIMPLES"
    
    def __init__(self, config: Optional[Configuracoes] = None):
        """
        Inicializa o carregador de descrições.
        
        Args:
            config: Configurações do sistema (opcional)
        """
        self.config = config or Configuracoes()
        self.dados: Dict[str, DadosProduto] = {}
        self._carregado: bool = False
        self._data_carregamento: Optional[datetime] = None
        self._erros: List[str] = []
        self._linha_cabecalho: int = 1

    def _prefixos_para_remover(self) -> List[str]:
        prefixos = list(self.config.mapa_prefixo_conta.keys())
        mapa_full = getattr(self.config, "mapa_prefixo_conta_full", {}) or {}
        for prefixo in mapa_full.keys():
            if prefixo not in prefixos:
                prefixos.append(prefixo)
        return prefixos
    
    @property
    def esta_carregado(self) -> bool:
        """Verifica se os dados foram carregados."""
        return self._carregado
    
    @property
    def quantidade_registros(self) -> int:
        """Retorna a quantidade de registros carregados."""
        return len(self.dados)
    
    @property
    def erros(self) -> List[str]:
        """Retorna lista de erros ocorridos durante carregamento."""
        return self._erros.copy()
    
    def arquivo_existe(self) -> bool:
        """Verifica se o arquivo de descrição existe."""
        return os.path.exists(self.config.arquivo_descricao)
    
    def carregar(self, forcar_recarga: bool = False) -> bool:
        """
        Carrega a base de dados de descrições.
        
        Args:
            forcar_recarga: Se True, recarrega mesmo se já carregado
            
        Returns:
            True se carregamento foi bem-sucedido
            
        Raises:
            FileNotFoundError: Se arquivo não existe
            Exception: Para outros erros de leitura
        """
        if self._carregado and not forcar_recarga:
            return True
        
        self._erros.clear()
        
        if not self.arquivo_existe():
            self._erros.append(f"Arquivo não encontrado: {self.config.arquivo_descricao}")
            raise FileNotFoundError(
                f"Base de Descrição não encontrada: {self.config.arquivo_descricao}"
            )
        
        try:
            # Abre a planilha
            workbook = openpyxl.load_workbook(
                self.config.arquivo_descricao,
                data_only=True
            )
            
            # Seleciona a aba (preferência para "PRODUTO SIMPLES")
            if self.ABA_PREFERIDA in workbook.sheetnames:
                planilha = workbook[self.ABA_PREFERIDA]
            else:
                planilha = workbook.active
            
            # 1. Encontra a linha do cabeçalho
            linha_cabecalho, dados_cabecalho = self._encontrar_cabecalho(planilha)
            self._linha_cabecalho = linha_cabecalho
            
            # 2. Mapeia as colunas
            indices = self._mapear_colunas(dados_cabecalho)

            # Fallback: se não achou coluna de SKU pelo mapa, usa coluna A (índice 0)
            if indices.get('sku') is None:
                indices['sku'] = 0
            
            # 3. Carrega os dados
            self.dados.clear()
            linha_inicio_dados = linha_cabecalho + 1
            
            for linha in planilha.iter_rows(min_row=linha_inicio_dados, values_only=True):
                dados_produto = self._processar_linha(linha, indices)
                if dados_produto:
                    self.dados[dados_produto.sku_base] = dados_produto
            
            self._carregado = True
            self._data_carregamento = datetime.now()
            
            return True
            
        except Exception as e:
            self._erros.append(f"Erro ao carregar descrições: {str(e)}")
            raise
    
    def _encontrar_cabecalho(self, planilha) -> Tuple[int, tuple]:
        """
        Encontra a linha do cabeçalho na planilha.
        
        Procura nas primeiras 10 linhas por uma linha que contenha valores.
        
        Args:
            planilha: Objeto worksheet do openpyxl
            
        Returns:
            Tupla (numero_linha, dados_cabecalho)
            
        Raises:
            ValueError: Se não encontrar cabeçalho válido
        """
        for numero_linha in range(1, 11):
            linha = next(planilha.iter_rows(
                min_row=numero_linha, 
                max_row=numero_linha, 
                values_only=True
            ))
            
            # Verifica se a linha tem conteúdo significativo
            if any(celula for celula in linha if celula is not None and str(celula).strip() != ""):
                return numero_linha, linha
        
        raise ValueError("Cabeçalho não encontrado nas primeiras 10 linhas")
    
    def _mapear_colunas(self, cabecalho: tuple) -> Dict[str, Optional[int]]:
        """
        Mapeia os nomes das colunas para seus índices.
        
        Usa o mapeamento definido em config.mapa_colunas_descricao para
        encontrar as colunas mesmo com nomes diferentes.
        
        Args:
            cabecalho: Tupla com os valores da linha de cabeçalho
            
        Returns:
            Dicionário {campo: indice} ou {campo: None}
        """
        # Normaliza o cabeçalho
        cabecalho_normalizado = [
            Utilitarios.normalizar_texto(h) if h else "" 
            for h in cabecalho
        ]
        
        indices = {}
        
        for campo, nomes_possiveis in self.config.mapa_colunas_descricao.items():
            indices[campo] = None
            
            for nome in nomes_possiveis:
                nome_normalizado = Utilitarios.normalizar_texto(nome)
                if nome_normalizado in cabecalho_normalizado:
                    indices[campo] = cabecalho_normalizado.index(nome_normalizado)
                    break
        
        return indices
    
    def _processar_linha(self, linha: tuple, indices: Dict[str, Optional[int]]) -> Optional[DadosProduto]:
        """
        Processa uma linha da planilha e cria um objeto DadosProduto.
        
        Args:
            linha: Tupla com os valores da linha
            indices: Mapeamento de campos para índices
            
        Returns:
            Objeto DadosProduto ou None se linha inválida
        """
        indice_sku = indices.get('sku')
        
        if indice_sku is None or indice_sku >= len(linha):
            return None
        
        sku_valor = linha[indice_sku]
        if not sku_valor:
            return None
        
        # Trata o SKU
        sku_base = Utilitarios.tratar_sku(
            sku_valor,
            self._prefixos_para_remover()
        )
        
        if not sku_base:
            return None
        
        # Função auxiliar para obter valor de uma coluna
        def obter_valor(campo: str) -> str:
            indice = indices.get(campo)
            if indice is not None and indice < len(linha):
                valor = linha[indice]
                return str(valor).strip() if valor is not None else ""
            return ""
        
        # Coleta os tópicos
        topicos = []
        for i in range(1, 6):
            topico = obter_valor(f'topico{i}')
            if topico:
                topicos.append(topico)
        
        return DadosProduto(
            sku_base=sku_base,
            titulo=obter_valor('titulo'),
            descricao=obter_valor('desc'),
            modelo=obter_valor('modelo'),
            ean=obter_valor('ean'),
            peso=obter_valor('peso'),
            comprimento=obter_valor('comp'),
            largura=obter_valor('larg'),
            altura=obter_valor('alt'),
            topicos=topicos
        )
    
    def obter_produto(self, sku: str) -> Optional[DadosProduto]:
        """
        Obtém os dados de um produto pelo SKU.
        
        Suporta busca direta e cálculo de Kits.
        
        Args:
            sku: SKU do produto (pode ser Kit)
            
        Returns:
            Objeto DadosProduto ou None
        """
        if not self._carregado:
            return None
        
        sku_base = Utilitarios.tratar_sku(
            sku,
            self._prefixos_para_remover()
        )
        
        if not sku_base:
            return None
        
        # Busca direta
        if sku_base in self.dados:
            return self.dados[sku_base]
        
        # Tenta calcular como Kit
        dados_kit = self._calcular_kit(sku_base)
        if dados_kit:
            return dados_kit
        
        return None
    
    def obter_produto_como_dict(self, sku: str) -> Optional[Dict[str, Any]]:
        """
        Obtém os dados de um produto como dicionário (compatibilidade).
        
        Args:
            sku: SKU do produto
            
        Returns:
            Dicionário com os dados ou None
        """
        produto = self.obter_produto(sku)
        if produto:
            return produto.para_dicionario()
        return None
    
    def _calcular_kit(self, sku_kit: str) -> Optional[DadosProduto]:
        """
        Calcula os dados de um Kit baseado nos produtos individuais.
        
        Regras de cálculo:
          - Peso e Altura: soma (ou base * quantidade)
          - Comprimento e Largura: maior valor entre componentes
        
        Formatos suportados:
          - Kx-SKU  (ex: K2-2999)  → x unidades do mesmo produto
          - K-A-B   (ex: K-2999-6392) → produtos diferentes
        """
        resultado = self._calcular_kit_quantidade(sku_kit)
        if resultado:
            return resultado
        
        return self._calcular_kit_misto(sku_kit)
    
    def _calcular_kit_quantidade(self, sku_kit: str) -> Optional[DadosProduto]:
        """
        Kit com multiplicador de quantidade.
        
        Formatos:
          - Kx-SKU       (ex: K2-2999)       → x unidades do mesmo produto
          - Kx-A-B       (ex: K5-6384-6392)   → x kits de produtos A+B
        """
        match = re.match(r"^K(\d+)-(.+)$", sku_kit, re.IGNORECASE)
        if not match:
            return None
        
        quantidade = int(match.group(1))
        resto = match.group(2).strip().upper()
        
        # Remove sufixo -V se presente
        if "-V" in resto:
            resto = resto.split("-V")[0]
        
        # Tenta como produto único (ex: K2-2999)
        produto_base = self.dados.get(resto)
        if produto_base:
            peso = Utilitarios.valor_para_float(produto_base.peso) * quantidade
            altura = Utilitarios.valor_para_float(produto_base.altura) * quantidade
            comprimento = Utilitarios.valor_para_float(produto_base.comprimento)
            largura = Utilitarios.valor_para_float(produto_base.largura)
            
            return DadosProduto(
                sku_base=sku_kit,
                titulo=produto_base.titulo,
                descricao=produto_base.descricao,
                modelo=produto_base.modelo,
                ean=produto_base.ean,
                peso=str(peso),
                comprimento=str(comprimento),
                largura=str(largura),
                altura=str(altura),
                topicos=produto_base.topicos.copy()
            )
        
        # Se contém '-', é kit misto com multiplicador (ex: K5-6384-6392)
        if '-' in resto:
            partes = resto.split('-')
            primeiro_produto = None
            peso_base = 0.0
            altura_base = 0.0
            comprimento_max = 0.0
            largura_max = 0.0
            
            for sku_parte in partes:
                sku_parte = sku_parte.strip().upper()
                if not sku_parte:
                    continue
                produto = self.dados.get(sku_parte)
                if produto:
                    if primeiro_produto is None:
                        primeiro_produto = produto
                    peso_base += Utilitarios.valor_para_float(produto.peso)
                    altura_base += Utilitarios.valor_para_float(produto.altura)
                    comprimento_max = max(
                        comprimento_max,
                        Utilitarios.valor_para_float(produto.comprimento)
                    )
                    largura_max = max(
                        largura_max,
                        Utilitarios.valor_para_float(produto.largura)
                    )
            
            if primeiro_produto:
                return DadosProduto(
                    sku_base=sku_kit,
                    titulo=primeiro_produto.titulo,
                    descricao=primeiro_produto.descricao,
                    modelo=primeiro_produto.modelo,
                    ean=primeiro_produto.ean,
                    peso=str(peso_base * quantidade),
                    comprimento=str(comprimento_max),
                    largura=str(largura_max),
                    altura=str(altura_base * quantidade),
                    topicos=primeiro_produto.topicos.copy()
                )
        
        return None
    
    def _calcular_kit_misto(self, sku_kit: str) -> Optional[DadosProduto]:
        """Kit K-A-B-C: produtos diferentes combinados."""
        if not sku_kit.upper().startswith("K-"):
            return None
        
        partes = sku_kit.split('-')[1:]  # Remove o "K" inicial
        
        primeiro_produto = None
        peso_total = 0.0
        altura_total = 0.0
        comprimento_max = 0.0
        largura_max = 0.0
        
        for sku_parte in partes:
            sku_parte = sku_parte.strip().upper()
            if not sku_parte:
                continue
            
            produto = self.dados.get(sku_parte)
            if produto:
                if primeiro_produto is None:
                    primeiro_produto = produto
                peso_total += Utilitarios.valor_para_float(produto.peso)
                altura_total += Utilitarios.valor_para_float(produto.altura)
                comprimento_max = max(
                    comprimento_max,
                    Utilitarios.valor_para_float(produto.comprimento)
                )
                largura_max = max(
                    largura_max,
                    Utilitarios.valor_para_float(produto.largura)
                )
        
        if not primeiro_produto:
            return None
        
        return DadosProduto(
            sku_base=sku_kit,
            titulo=primeiro_produto.titulo,
            descricao=primeiro_produto.descricao,
            modelo=primeiro_produto.modelo,
            ean=primeiro_produto.ean,
            peso=str(peso_total),
            comprimento=str(comprimento_max),
            largura=str(largura_max),
            altura=str(altura_total),
            topicos=primeiro_produto.topicos.copy()
        )
    
    def limpar_cache(self) -> None:
        """Limpa os dados carregados do cache."""
        self.dados.clear()
        self._carregado = False
        self._data_carregamento = None
        self._erros.clear()
