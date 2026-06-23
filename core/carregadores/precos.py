# ==============================================================================
# CARREGADOR DE PREÇOS
# ==============================================================================
# Classe responsável por carregar e gerenciar a base de dados de preços.
#
# Funcionalidades:
#   - Carrega planilha de precificação
#   - Mapeia preços por conta (baseado no prefixo do SKU)
#   - Calcula preço com base no prefixo do SKU
# ==============================================================================

import os
import openpyxl
from typing import Dict, Optional, Any, List
from dataclasses import dataclass, field
from datetime import datetime

from ..utils import Utilitarios
from ..config import Configuracoes


@dataclass
class DadosPreco:
    """
    Estrutura para armazenar os preços de um produto por conta.
    
    Attributes:
        sku_base: SKU base do produto (sem prefixo/sufixo)
        precos_por_conta: Dicionário com preço por prefixo de conta
        preco_padrao: Preço padrão quando não há preço específico
    """
    sku_base: str
    precos_por_conta: Dict[str, Any] = field(default_factory=dict)
    preco_padrao: Optional[Any] = None


class CarregadorPrecos:
    """
    Carrega e gerencia a base de dados de preços.
    
    Esta classe é responsável por:
    - Carregar a planilha de precificação
    - Mapear dinamicamente as colunas por nome
    - Retornar o preço correto baseado no prefixo do SKU
    
    Attributes:
        config: Instância de Configuracoes com os parâmetros
        dados: Dicionário com os dados carregados
        _carregado: Flag indicando se os dados foram carregados
    """
    
    def __init__(self, config: Optional[Configuracoes] = None, *,
                 arquivo: Optional[str] = None,
                 aba: Optional[str] = None,
                 mapa_prefixo: Optional[Dict[str, str]] = None):
        """
        Inicializa o carregador de preços.

        Args:
            config: Configurações do sistema (opcional, usa padrão se não fornecido)
            arquivo: Caminho do arquivo de precificação (override; default =
                config.arquivo_precificacao). Usado pela modalidade FULL para
                apontar para a "Precificacao Amazon - Full".
            aba: Nome da aba a carregar (override; default = aba ativa). FULL usa "CLA".
            mapa_prefixo: Mapa {prefixo_sku: nome_coluna_conta} (override; default =
                config.mapa_prefixo_conta). FULL usa os prefixos CLA.
        """
        self.config = config or Configuracoes()
        self._arquivo_override = arquivo
        self._aba_override = aba
        self._mapa_prefixo_override = mapa_prefixo
        self.dados: Dict[str, DadosPreco] = {}
        self._carregado: bool = False
        self._data_carregamento: Optional[datetime] = None
        self._erros: List[str] = []

    @property
    def _arquivo(self) -> str:
        """Arquivo de precificação efetivo (override ou padrão da config)."""
        return self._arquivo_override or self.config.arquivo_precificacao

    @property
    def _mapa_prefixo(self) -> Dict[str, str]:
        """Mapa prefixo→conta efetivo (override ou padrão da config)."""
        return self._mapa_prefixo_override or self.config.mapa_prefixo_conta
    
    @property
    def esta_carregado(self) -> bool:
        """Verifica se os dados foram carregados."""
        return self._carregado
    
    @property
    def quantidade_registros(self) -> int:
        """Retorna a quantidade de registros carregados."""
        return len(self.dados)
    
    @property
    def erros(self) -> List[str]:
        """Retorna lista de erros ocorridos durante carregamento."""
        return self._erros.copy()
    
    def arquivo_existe(self) -> bool:
        """Verifica se o arquivo de precificação existe."""
        return os.path.exists(self._arquivo)
    
    def carregar(self, forcar_recarga: bool = False) -> bool:
        """
        Carrega a base de dados de preços.
        
        Args:
            forcar_recarga: Se True, recarrega mesmo se já carregado
            
        Returns:
            True se carregamento foi bem-sucedido
            
        Raises:
            FileNotFoundError: Se arquivo não existe
            Exception: Para outros erros de leitura
        """
        # Evita recarga desnecessária
        if self._carregado and not forcar_recarga:
            return True
        
        self._erros.clear()
        
        # Verifica existência do arquivo
        if not self.arquivo_existe():
            self._erros.append(f"Arquivo não encontrado: {self._arquivo}")
            raise FileNotFoundError(
                f"Base de Precificação não encontrada: {self._arquivo}"
            )

        try:
            # Abre a planilha
            workbook = openpyxl.load_workbook(
                self._arquivo,
                data_only=True
            )
            # Aba específica (FULL usa "CLA") ou a aba ativa por padrão
            if self._aba_override and self._aba_override in workbook.sheetnames:
                planilha = workbook[self._aba_override]
            else:
                planilha = workbook.active
            
            # 1. Lê o cabeçalho (linha 1) e mapeia colunas
            linha_cabecalho = next(planilha.iter_rows(min_row=1, max_row=1, values_only=True))
            mapa_colunas = self._mapear_colunas_cabecalho(linha_cabecalho)
            
            # 2. Identifica coluna do SKU (chave de busca)
            indice_sku = self._encontrar_coluna_sku(mapa_colunas)
            
            # 3. Identifica coluna do Preço Padrão
            indice_preco_padrao = self._encontrar_coluna_preco_padrao(mapa_colunas)
            
            # 4. Mapeia colunas das contas (prefixos)
            indices_contas = self._mapear_colunas_contas(mapa_colunas)
            
            if indice_preco_padrao is not None:
                indices_contas['PADRAO'] = indice_preco_padrao
            
            # 5. Carrega os dados
            self.dados.clear()
            
            for linha in planilha.iter_rows(min_row=2, values_only=True):
                if indice_sku < len(linha) and linha[indice_sku]:
                    sku_base = Utilitarios.tratar_sku(
                        linha[indice_sku],
                        list(self._mapa_prefixo.keys())
                    )
                    
                    if sku_base:
                        precos_conta = {}
                        for prefixo, indice in indices_contas.items():
                            if indice < len(linha) and linha[indice]:
                                precos_conta[prefixo] = linha[indice]
                        
                        self.dados[sku_base] = DadosPreco(
                            sku_base=sku_base,
                            precos_por_conta=precos_conta,
                            preco_padrao=precos_conta.get('PADRAO')
                        )
            
            self._carregado = True
            self._data_carregamento = datetime.now()
            
            return True
            
        except Exception as e:
            self._erros.append(f"Erro ao carregar preços: {str(e)}")
            raise
    
    def _mapear_colunas_cabecalho(self, linha_cabecalho: tuple) -> Dict[str, int]:
        """
        Mapeia os nomes das colunas para seus índices.
        
        Args:
            linha_cabecalho: Tupla com os valores da linha de cabeçalho
            
        Returns:
            Dicionário {nome_normalizado: indice}
        """
        mapa = {}
        for indice, valor in enumerate(linha_cabecalho):
            if valor:
                chave_normalizada = Utilitarios.normalizar_texto(str(valor))
                mapa[chave_normalizada] = indice
        return mapa
    
    def _nomes_para_chave(self, chave: str, padrao: List[str]) -> List[str]:
        """
        Retorna a lista de sinônimos configurada para uma chave lógica do
        mapa de colunas da Precificação. Se a configuração estiver vazia,
        usa os padrões fornecidos.
        """
        mapa = getattr(self.config, 'mapa_colunas_precificacao', {}) or {}
        nomes = mapa.get(chave) or []
        if not nomes:
            return list(padrao)
        return list(nomes)

    def _encontrar_coluna_sku(self, mapa_colunas: Dict[str, int]) -> int:
        """
        Encontra o índice da coluna de SKU.

        Os sinônimos vêm do mapa configurável `mapa_colunas_precificacao['sku']`.

        Returns:
            Índice da coluna de SKU (0 como fallback)
        """
        nomes_possiveis = self._nomes_para_chave(
            'sku', ['SKU', 'Chave', 'Código', 'Item']
        )

        for nome in nomes_possiveis:
            nome_normalizado = Utilitarios.normalizar_texto(nome)
            if nome_normalizado in mapa_colunas:
                return mapa_colunas[nome_normalizado]

        # Fallback: assume primeira coluna
        return 0

    def _encontrar_coluna_preco_padrao(self, mapa_colunas: Dict[str, int]) -> Optional[int]:
        """
        Encontra o índice da coluna de preço padrão.

        Os sinônimos vêm do mapa configurável
        `mapa_colunas_precificacao['preco_padrao']`.
        """
        nomes_possiveis = self._nomes_para_chave(
            'preco_padrao', ['Padrão', 'Preço Padrão', 'Standard']
        )

        for nome in nomes_possiveis:
            nome_normalizado = Utilitarios.normalizar_texto(nome)
            if nome_normalizado in mapa_colunas:
                return mapa_colunas[nome_normalizado]

        return None
    
    def _mapear_colunas_contas(self, mapa_colunas: Dict[str, int]) -> Dict[str, int]:
        """
        Mapeia as colunas de cada conta (prefixo) para seus índices.
        
        Args:
            mapa_colunas: Mapeamento de colunas
            
        Returns:
            Dicionário {prefixo: indice_coluna}
        """
        indices = {}

        for prefixo, nome_conta in self._mapa_prefixo.items():
            chave_normalizada = Utilitarios.normalizar_texto(nome_conta)
            if chave_normalizada in mapa_colunas:
                indices[prefixo] = mapa_colunas[chave_normalizada]
        
        return indices
    
    def obter_preco(self, sku_completo: str) -> Optional[str]:
        """
        Obtém o preço formatado para um SKU específico.
        
        O preço é determinado pelo prefixo do SKU. Se não houver preço
        específico para a conta, usa o preço padrão.
        
        Args:
            sku_completo: SKU completo com prefixo (ex: NOGO-ABC123)
            
        Returns:
            Preço formatado como string (ex: "99.90") ou None
        """
        if not self._carregado:
            return None
        
        # Trata o SKU para obter a chave de busca
        sku_base = Utilitarios.tratar_sku(
            sku_completo,
            list(self._mapa_prefixo.keys())
        )

        if not sku_base or sku_base not in self.dados:
            return None

        dados_preco = self.dados[sku_base]

        # Identifica o prefixo do SKU original
        sku_upper = str(sku_completo).upper()
        prefixo_encontrado = None

        for prefixo in self._mapa_prefixo.keys():
            if sku_upper.startswith(prefixo):
                prefixo_encontrado = prefixo
                break
        
        # Obtém o preço da conta ou o padrão
        if prefixo_encontrado:
            preco = dados_preco.precos_por_conta.get(prefixo_encontrado)
        else:
            preco = None
        
        if preco is None:
            preco = dados_preco.preco_padrao or dados_preco.precos_por_conta.get('PADRAO')
        
        if preco is not None:
            return Utilitarios.formatar_decimal(preco)
        
        return None
    
    def limpar_cache(self) -> None:
        """Limpa os dados carregados do cache."""
        self.dados.clear()
        self._carregado = False
        self._data_carregamento = None
        self._erros.clear()
