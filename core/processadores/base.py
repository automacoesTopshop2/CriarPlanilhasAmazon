# ==============================================================================
# PROCESSADOR BASE (CLASSE ABSTRATA)
# ==============================================================================
# Classe base abstrata para todos os processadores de templates.
#
# Esta classe define a interface comum e métodos reutilizáveis para:
#   - Processamento de planilhas Amazon
#   - Mapeamento dinâmico de colunas
#   - Escrita de dados em templates
#   - Geração de logs e erros
#
# Para criar novos processadores (ex: Brinquedos, Suplementos), basta
# herdar desta classe e implementar os métodos abstratos.
#
# NOTA sobre preservação de estilos:
#   O openpyxl pode corromper estilos de células com cores indexadas
#   ao modificar valores. A opção 'preservar_estilo' nos métodos de
#   escrita mitiga esse problema copiando o estilo antes da escrita.
# ==============================================================================

import io
import openpyxl
from copy import copy
from abc import ABC, abstractmethod
from typing import Dict, List, Optional, Any, Tuple, Callable
from dataclasses import dataclass, field
from datetime import datetime

from ..config import Configuracoes
from ..utils import Utilitarios, LogErro
from ..mapeadores import MapeadorColunas
from ..carregadores import CarregadorPrecos, CarregadorDescricao, CarregadorDescricaoAPI


@dataclass
class ResultadoProcessamento:
    """
    Estrutura para armazenar o resultado do processamento.
    
    Attributes:
        sucesso: Se processamento foi bem-sucedido
        arquivo_saida: Buffer com o arquivo gerado
        total_processados: Quantidade de linhas processadas
        total_erros: Quantidade de erros encontrados
        total_avisos: Quantidade de avisos gerados
        logs: Lista de logs de erro/aviso
        tempo_processamento: Duração do processamento
    """
    sucesso: bool = False
    arquivo_saida: Optional[io.BytesIO] = None
    nome_arquivo: str = ""
    total_processados: int = 0
    total_erros: int = 0
    total_avisos: int = 0
    logs: List[LogErro] = field(default_factory=list)
    tempo_processamento: float = 0.0
    mensagem: str = ""


class ProcessadorBase(ABC):
    """
    Classe base abstrata para processadores de templates Amazon.
    
    Esta classe fornece:
    - Interface comum para todos os processadores
    - Métodos utilitários reutilizáveis
    - Gerenciamento de logs e erros
    - Configuração flexível por categoria
    
    Para criar um novo processador, herde desta classe e implemente:
    - processar(): Lógica principal de processamento
    - _obter_nome_arquivo_saida(): Nome do arquivo gerado
    
    Attributes:
        config: Configurações do sistema
        carregador_precos: Instância do carregador de preços
        carregador_descricao: Instância do carregador de descrições
        mapeador: Instância do mapeador de colunas
        logs: Lista de logs gerados durante processamento
    """
    
    # -------------------------------------------------------------------------
    # CONSTANTES (podem ser sobrescritas em subclasses)
    # -------------------------------------------------------------------------
    
    LINHA_CABECALHO_TEMPLATE: int = 4      # Linha com nomes das colunas no template
    LINHA_INICIO_DADOS_PADRAO: int = 8     # Linha inicial padrão para escrita
    ABA_TEMPLATE: str = "Modelo"           # Nome da aba principal do template
    
    def __init__(self, config: Optional[Configuracoes] = None):
        """
        Inicializa o processador base.
        
        Args:
            config: Configurações do sistema (opcional, usa padrão)
        """
        self.config = config or Configuracoes()
        self.carregador_precos = CarregadorPrecos(self.config)
        # Fonte da base de Descrição: API do AgentedeTitulos (quando
        # TITULOS_API_KEY está setada) ou a planilha DESCRIÇÃO.xlsx (padrão).
        if self.config.usar_api_descricao:
            self.carregador_descricao = CarregadorDescricaoAPI(self.config)
        else:
            self.carregador_descricao = CarregadorDescricao(self.config)
        self.mapeador = MapeadorColunas()
        self.logs: List[LogErro] = []
        
        # Estado interno
        self._workbook_saida = None
        self._planilha_saida = None
        self._linha_atual = self.LINHA_INICIO_DADOS_PADRAO
        self._callback_progresso: Optional[Callable[[float], None]] = None
    
    # -------------------------------------------------------------------------
    # MÉTODOS ABSTRATOS (devem ser implementados pelas subclasses)
    # -------------------------------------------------------------------------
    
    @abstractmethod
    def processar(self, arquivo_entrada: Any, arquivo_template: Any,
                  callback_status: Optional[Callable[[str], None]] = None,
                  callback_progresso: Optional[Callable[[float], None]] = None) -> ResultadoProcessamento:
        """
        Processa os dados e gera o arquivo de saída.
        
        Args:
            arquivo_entrada: Arquivo com os dados de entrada
            arquivo_template: Arquivo de template para preenchimento
            callback_status: Função para reportar status textual
            callback_progresso: Função para reportar progresso (0.0 a 1.0)
            
        Returns:
            Objeto ResultadoProcessamento com o resultado
        """
        pass
    
    @abstractmethod
    def _obter_nome_arquivo_saida(self) -> str:
        """
        Retorna o nome sugerido para o arquivo de saída.
        
        Returns:
            Nome do arquivo com extensão
        """
        pass
    
    # -------------------------------------------------------------------------
    # MÉTODOS DE SETUP (inicialização de recursos)
    # -------------------------------------------------------------------------
    
    def _carregar_bases_dados(self, callback_status: Optional[Callable] = None) -> bool:
        """
        Carrega as bases de dados de preços e descrições.
        
        Args:
            callback_status: Callback para reportar status
            
        Returns:
            True se carregamento bem-sucedido
        """
        try:
            if callback_status:
                callback_status("📚 Carregando base de Preços...")
            self.carregador_precos.carregar()
            
            if callback_status:
                callback_status("📚 Carregando base de Descrição...")
            self.carregador_descricao.carregar()
            
            return True
            
        except FileNotFoundError as e:
            self._adicionar_erro("SISTEMA", "Erro", str(e))
            raise
        except Exception as e:
            self._adicionar_erro("SISTEMA", "Erro", f"Falha ao carregar bases: {str(e)}", incluir_traceback=True)
            raise
    
    def _abrir_template(self, arquivo_template: Any, 
                        callback_status: Optional[Callable] = None) -> Tuple[Any, Any]:
        """
        Abre o arquivo de template para edição.
        
        Args:
            arquivo_template: Arquivo de template (file-like ou path)
            callback_status: Callback para reportar status
            
        Returns:
            Tupla (workbook, worksheet)
        """
        if callback_status:
            callback_status("📂 Abrindo template...")
        
        try:
            workbook = openpyxl.load_workbook(arquivo_template, keep_vba=True)
            
            # Tenta encontrar a aba do modelo
            if self.ABA_TEMPLATE in workbook.sheetnames:
                planilha = workbook[self.ABA_TEMPLATE]
            else:
                planilha = workbook.active
            
            self._workbook_saida = workbook
            self._planilha_saida = planilha
            
            return workbook, planilha
            
        except Exception as e:
            self._adicionar_erro("TEMPLATE", "Erro", f"Falha ao abrir template: {str(e)}", incluir_traceback=True)
            raise
    
    def _mapear_template(self, planilha, callback_status: Optional[Callable] = None) -> Dict[str, List[int]]:
        """
        Mapeia as colunas do template.
        
        A linha inicial de escrita usa LINHA_INICIO_DADOS_PADRAO (valor fixo)
        em vez de detecção automática, pois templates podem ter dados de
        exemplo que confundiriam a detecção.
        
        Args:
            planilha: Worksheet do template
            callback_status: Callback para reportar status
            
        Returns:
            Dicionário de mapeamento de colunas
        """
        if callback_status:
            callback_status("🗺️ Mapeando colunas do template...")
        
        # Mapeia colunas (sem detecção de primeira linha - usamos valor fixo)
        mapa = self.mapeador.mapear_template(
            planilha, 
            self.LINHA_CABECALHO_TEMPLATE,
            detectar_primeira_linha=False
        )
        
        # Usa linha inicial fixa (definida na classe)
        # SKU: LINHA_INICIO_DADOS_PADRAO = 8
        # ASIN: LINHA_INICIO_DADOS_PADRAO = 7
        self._linha_atual = self.LINHA_INICIO_DADOS_PADRAO
        
        return mapa
    
    # -------------------------------------------------------------------------
    # MÉTODOS DE ESCRITA (preenchimento do template)
    # -------------------------------------------------------------------------
    
    def _escrever_valor_celula(self, celula, valor: Any, preservar_estilo: bool = True) -> None:
        """
        Escreve um valor em uma célula, opcionalmente preservando o estilo.
        
        O openpyxl pode corromper estilos (especialmente cores indexadas)
        ao modificar células. Esta função preserva fill, font, border e
        alignment antes de escrever o valor.
        
        Args:
            celula: Objeto Cell do openpyxl
            valor: Valor a escrever
            preservar_estilo: Se True, copia e reaplica os estilos
        """
        if preservar_estilo:
            # Salva os estilos existentes ANTES de modificar
            try:
                estilo_preenchimento = copy(celula.fill) if celula.fill else None
                estilo_fonte = copy(celula.font) if celula.font else None
                estilo_borda = copy(celula.border) if celula.border else None
                estilo_alinhamento = copy(celula.alignment) if celula.alignment else None
            except Exception:
                # Se falhar ao copiar, apenas escreve sem preservar
                celula.value = valor
                return
            
            # Escreve o valor
            celula.value = valor
            
            # Reaplica os estilos
            try:
                if estilo_preenchimento:
                    celula.fill = estilo_preenchimento
                if estilo_fonte:
                    celula.font = estilo_fonte
                if estilo_borda:
                    celula.border = estilo_borda
                if estilo_alinhamento:
                    celula.alignment = estilo_alinhamento
            except Exception:
                pass  # Ignora erros de estilo
        else:
            celula.value = valor
    
    def _escrever_valor(self, linha: int, nome_coluna: str, valor: Any,
                       apenas_primeira: bool = False) -> bool:
        """
        Escreve um valor em uma coluna pelo nome, preservando estilos.
        
        Args:
            linha: Número da linha para escrita
            nome_coluna: Nome da coluna (será normalizado)
            valor: Valor a escrever
            apenas_primeira: Se True, escreve apenas na primeira ocorrência
            
        Returns:
            True se conseguiu escrever
        """
        if self._planilha_saida is None or valor is None:
            return False
        
        indices = self.mapeador.obter_indices(nome_coluna)
        
        if not indices:
            return False
        
        if apenas_primeira:
            celula = self._planilha_saida.cell(row=linha, column=indices[0])
            self._escrever_valor_celula(celula, valor, preservar_estilo=True)
        else:
            for indice in indices:
                celula = self._planilha_saida.cell(row=linha, column=indice)
                self._escrever_valor_celula(celula, valor, preservar_estilo=True)
        
        return True
    
    def _escrever_valores_fixos(self, linha: int) -> None:
        """
        Escreve os valores fixos configurados na linha especificada.
        
        Args:
            linha: Número da linha para escrita
        """
        for nome_coluna, valor in self.config.valores_fixos_padrao.items():
            # Verifica se deve preencher apenas a primeira ocorrência
            apenas_primeira = any(
                Utilitarios.normalizar_texto(col) == Utilitarios.normalizar_texto(nome_coluna)
                for col in self.config.colunas_apenas_primeira_ocorrencia
            )
            self._escrever_valor(linha, nome_coluna, valor, apenas_primeira)
    
    def _escrever_imagens(self, linha: int, sku_base: str, 
                          quantidade_imagens: int = 5) -> None:
        """
        Escreve as URLs de imagens no template.
        
        Args:
            linha: Número da linha para escrita
            sku_base: SKU base para construir URLs
            quantidade_imagens: Quantidade total de imagens (principal + extras)
        """
        if self._planilha_saida is None:
            return
        
        mapa_template = self.mapeador.mapa
        contador_imagem_extra = 2  # Começa em 02 (01 é principal)
        
        for nome_coluna, indices in mapa_template.items():
            # Imagem principal
            if any(k in nome_coluna for k in self.config.palavras_chave_imagem_principal):
                url = f"{self.config.url_base_imagens}/{sku_base}/{sku_base}_01.jpg"
                celula = self._planilha_saida.cell(row=linha, column=indices[0])
                self._escrever_valor_celula(celula, url)
                continue
            
            # Imagens secundárias
            eh_imagem_secundaria = (
                ('outro' in nome_coluna and 'imagem' in nome_coluna) or
                'other image' in nome_coluna or
                (('imagem' in nome_coluna or 'image' in nome_coluna) and
                 not any(k in nome_coluna for k in ['principal', 'main']))
            )
            
            if eh_imagem_secundaria:
                for indice in indices:
                    if contador_imagem_extra <= quantidade_imagens:
                        url = f"{self.config.url_base_imagens}/{sku_base}/{sku_base}_{contador_imagem_extra:02d}.jpg"
                        celula = self._planilha_saida.cell(row=linha, column=indice)
                        self._escrever_valor_celula(celula, url)
                        contador_imagem_extra += 1
    
    def _escrever_topicos(self, linha: int, topicos: List[str]) -> None:
        """
        Escreve os tópicos/bullets no template de forma sequencial.
        
        Args:
            linha: Número da linha para escrita
            topicos: Lista de tópicos a escrever
        """
        if self._planilha_saida is None or not topicos:
            return
        
        mapa_template = self.mapeador.mapa
        indice_topico = 0
        palavras_chave = ['topico', 'bullet', 'marcador', 'ponto de destaque']
        
        for nome_coluna, indices in mapa_template.items():
            if any(k in nome_coluna for k in palavras_chave):
                for indice in indices:
                    if indice_topico < len(topicos):
                        celula = self._planilha_saida.cell(row=linha, column=indice)
                        self._escrever_valor_celula(celula, topicos[indice_topico])
                        indice_topico += 1
    
    # -------------------------------------------------------------------------
    # MÉTODOS DE FINALIZAÇÃO
    # -------------------------------------------------------------------------
    
    def _gerar_arquivo_saida(self) -> io.BytesIO:
        """
        Gera o buffer de saída com o arquivo processado.
        
        Returns:
            BytesIO com o conteúdo do arquivo
        """
        buffer = io.BytesIO()
        self._workbook_saida.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _criar_resultado(self, sucesso: bool, arquivo: Optional[io.BytesIO] = None,
                        total_processados: int = 0, tempo: float = 0.0,
                        mensagem: str = "") -> ResultadoProcessamento:
        """
        Cria o objeto de resultado do processamento.
        
        Args:
            sucesso: Se processamento foi bem-sucedido
            arquivo: Buffer com arquivo gerado
            total_processados: Quantidade de linhas processadas
            tempo: Tempo de processamento em segundos
            mensagem: Mensagem de status
            
        Returns:
            Objeto ResultadoProcessamento
        """
        total_erros = sum(1 for log in self.logs if log.tipo == "Erro")
        total_avisos = sum(1 for log in self.logs if log.tipo == "Aviso")
        
        return ResultadoProcessamento(
            sucesso=sucesso,
            arquivo_saida=arquivo,
            nome_arquivo=self._obter_nome_arquivo_saida(),
            total_processados=total_processados,
            total_erros=total_erros,
            total_avisos=total_avisos,
            logs=self.logs.copy(),
            tempo_processamento=tempo,
            mensagem=mensagem
        )
    
    # -------------------------------------------------------------------------
    # MÉTODOS DE LOG
    # -------------------------------------------------------------------------
    
    def _adicionar_erro(self, sku: str, tipo: str, mensagem: str,
                       linha: Optional[int] = None,
                       incluir_traceback: bool = False) -> None:
        """
        Adiciona um log de erro/aviso.
        
        Args:
            sku: SKU relacionado
            tipo: Tipo ('Erro', 'Aviso', 'Info')
            mensagem: Descrição do problema
            linha: Número da linha (opcional)
            incluir_traceback: Se deve capturar stack trace
        """
        log = Utilitarios.criar_log_erro(
            sku=sku,
            tipo=tipo,
            mensagem=mensagem,
            linha=linha,
            incluir_traceback=incluir_traceback
        )
        self.logs.append(log)
    
    def limpar_logs(self) -> None:
        """Limpa a lista de logs."""
        self.logs.clear()
    
    # -------------------------------------------------------------------------
    # MÉTODOS AUXILIARES
    # -------------------------------------------------------------------------
    
    def _verificar_deve_preencher_apenas_primeira(self, nome_coluna: str) -> bool:
        """
        Verifica se uma coluna deve ser preenchida apenas na primeira ocorrência.
        
        Args:
            nome_coluna: Nome da coluna normalizado
            
        Returns:
            True se deve preencher apenas a primeira
        """
        nome_normalizado = Utilitarios.normalizar_texto(nome_coluna)
        
        for coluna_restrita in self.config.colunas_apenas_primeira_ocorrencia:
            if Utilitarios.normalizar_texto(coluna_restrita) == nome_normalizado:
                return True
        
        return False
