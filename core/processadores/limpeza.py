# ==============================================================================
# PROCESSADOR LIMPEZA
# ==============================================================================
# Processador para limpeza de textos em planilhas Amazon.
#
# Este processador:
#   - Remove termos indesejados
#   - Substitui termos por alternativas
#   - Normaliza espaços em branco
#   - Opera em colunas de texto (título, descrição, tópicos)
# ==============================================================================

import io
import os
import re
import time
import openpyxl
import traceback
from typing import Dict, List, Optional, Any, Callable, Tuple
from datetime import datetime

from .base import ProcessadorBase, ResultadoProcessamento
from ..config import Configuracoes
from ..utils import Utilitarios


class ProcessadorLimpeza(ProcessadorBase):
    """
    Processador para limpeza de textos em planilhas.
    
    Este processador é especializado em:
    - Remover termos configurados
    - Substituir termos por alternativas
    - Normalizar espaços (múltiplos espaços -> um)
    - Operar em colunas de texto identificadas automaticamente
    
    Herda de ProcessadorBase e implementa a lógica específica
    para limpeza de textos.
    """
    
    def __init__(self, config: Optional[Configuracoes] = None):
        """
        Inicializa o processador de limpeza.
        
        Args:
            config: Configurações do sistema (opcional)
        """
        super().__init__(config)
        self._termos_remover: List[str] = []
        self._termos_substituir: Dict[str, str] = {}
    
    def _obter_nome_arquivo_saida(self) -> str:
        """Retorna o nome do arquivo de saída."""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        return f"LIMPA_{timestamp}.xlsm"
    
    def carregar_termos(self) -> Tuple[List[str], Dict[str, str]]:
        """
        Carrega os termos de remoção e substituição dos arquivos.
        
        Returns:
            Tupla (lista_remover, dict_substituir)
        """
        arquivo_remover = self.config.arquivo_remover
        arquivo_substituir = self.config.arquivo_substituir
        
        # Garante que os arquivos existam
        Utilitarios.garantir_arquivos_txt(arquivo_remover, arquivo_substituir)
        
        # Carrega termos a remover
        self._termos_remover = []
        try:
            with open(arquivo_remover, 'r', encoding='utf-8', errors='ignore') as f:
                self._termos_remover = [linha.strip() for linha in f if linha.strip()]
        except Exception:
            pass
        
        # Carrega termos a substituir
        self._termos_substituir = {}
        try:
            with open(arquivo_substituir, 'r', encoding='utf-8', errors='ignore') as f:
                for linha in f:
                    if '=>' in linha:
                        partes = linha.strip().split('=>', 1)
                        if len(partes) == 2:
                            antigo, novo = partes[0].strip(), partes[1].strip()
                            if antigo:
                                self._termos_substituir[antigo] = novo
        except Exception:
            pass
        
        return self._termos_remover, self._termos_substituir
    
    def salvar_termo_remover(self, termo: str) -> bool:
        """
        Salva um novo termo a remover no arquivo.
        
        Args:
            termo: Termo a adicionar
            
        Returns:
            True se salvou com sucesso
        """
        try:
            with open(self.config.arquivo_remover, 'a', encoding='utf-8') as f:
                f.write(f"{termo}\n")
            return True
        except Exception:
            return False
    
    def salvar_termo_substituir(self, antigo: str, novo: str) -> bool:
        """
        Salva um novo par de substituição no arquivo.

        Args:
            antigo: Termo a ser substituído
            novo: Termo substituto

        Returns:
            True se salvou com sucesso
        """
        try:
            with open(self.config.arquivo_substituir, 'a', encoding='utf-8') as f:
                f.write(f"{antigo}=>{novo}\n")
            return True
        except Exception:
            return False

    def sobrescrever_termos_remover(self, termos: List[str]) -> bool:
        """
        Sobrescreve o arquivo de termos a remover com a lista informada.

        Args:
            termos: Lista completa de termos.

        Returns:
            True se salvou com sucesso.
        """
        try:
            Utilitarios.garantir_arquivos_txt(
                self.config.arquivo_remover, self.config.arquivo_substituir
            )
            with open(self.config.arquivo_remover, 'w', encoding='utf-8') as f:
                for termo in termos:
                    termo = (termo or "").strip()
                    if termo:
                        f.write(f"{termo}\n")
            return True
        except Exception:
            return False

    def sobrescrever_termos_substituir(self, dicio: Dict[str, str]) -> bool:
        """
        Sobrescreve o arquivo de termos a substituir.

        Args:
            dicio: Dicionário {antigo: novo}.

        Returns:
            True se salvou com sucesso.
        """
        try:
            Utilitarios.garantir_arquivos_txt(
                self.config.arquivo_remover, self.config.arquivo_substituir
            )
            with open(self.config.arquivo_substituir, 'w', encoding='utf-8') as f:
                for antigo, novo in dicio.items():
                    antigo = (antigo or "").strip()
                    novo = (novo or "").strip()
                    if antigo:
                        f.write(f"{antigo}=>{novo}\n")
            return True
        except Exception:
            return False
    
    def processar(self, arquivo_entrada: Any, arquivo_template: Any = None,
                  callback_status: Optional[Callable[[str], None]] = None,
                  callback_progresso: Optional[Callable[[float], None]] = None) -> ResultadoProcessamento:
        """
        Processa a limpeza de textos na planilha.
        
        Args:
            arquivo_entrada: Planilha a ser limpa
            arquivo_template: Não utilizado (mantido por compatibilidade)
            callback_status: Função para reportar status textual
            callback_progresso: Função para reportar progresso (0.0 a 1.0)
            
        Returns:
            Objeto ResultadoProcessamento com o resultado
        """
        inicio = time.time()
        self.limpar_logs()
        
        def status(msg: str):
            if callback_status:
                callback_status(msg)
        
        def progresso(valor: float):
            if callback_progresso:
                callback_progresso(valor)
        
        try:
            # 1. Carregar termos
            status("📋 Carregando termos de limpeza...")
            self.carregar_termos()
            
            # 2. Abrir planilha
            status("📂 Abrindo planilha para limpeza...")
            
            wb = openpyxl.load_workbook(arquivo_entrada, keep_vba=True)
            
            # Verifica se tem a aba Modelo
            if self.ABA_TEMPLATE not in wb.sheetnames:
                return self._criar_resultado(
                    sucesso=False,
                    tempo=time.time() - inicio,
                    mensagem=f"❌ Aba '{self.ABA_TEMPLATE}' não encontrada na planilha."
                )
            
            ws = wb[self.ABA_TEMPLATE]
            self._workbook_saida = wb
            self._planilha_saida = ws
            
            # 3. Mapear colunas
            status("🗺️ Identificando colunas de texto...")
            mapa_template = self.mapeador.mapear_template(ws, self.LINHA_CABECALHO_TEMPLATE)
            
            # Identifica colunas de texto
            colunas_texto = self._identificar_colunas_texto(mapa_template)
            
            if not colunas_texto:
                return self._criar_resultado(
                    sucesso=False,
                    tempo=time.time() - inicio,
                    mensagem="⚠️ Nenhuma coluna de texto identificada automaticamente."
                )
            
            # 4. Executar limpeza
            status("🧹 Executando limpeza...")
            
            celulas_modificadas = 0
            linha_inicio = self.mapeador.primeira_linha_dados
            
            # Conta total de linhas para progresso
            total_linhas = ws.max_row - linha_inicio + 1
            
            for idx_linha, linha in enumerate(ws.iter_rows(min_row=linha_inicio)):
                progresso((idx_linha + 1) / total_linhas)
                
                for celula in linha:
                    if celula.column not in colunas_texto:
                        continue
                    
                    if not isinstance(celula.value, str):
                        continue
                    
                    texto_original = celula.value
                    texto_limpo = self._limpar_texto(texto_original)
                    
                    if texto_limpo != texto_original:
                        # Usa método herdado para preservar estilos
                        self._escrever_valor_celula(celula, texto_limpo)
                        celulas_modificadas += 1
            
            # 5. Gerar arquivo de saída
            status("💾 Gerando arquivo limpo...")
            arquivo_saida = self._gerar_arquivo_saida()
            
            tempo_total = time.time() - inicio
            
            return self._criar_resultado(
                sucesso=True,
                arquivo=arquivo_saida,
                total_processados=celulas_modificadas,
                tempo=tempo_total,
                mensagem=f"✅ Limpeza concluída! {celulas_modificadas} células foram higienizadas."
            )
            
        except Exception as e:
            self._adicionar_erro(
                sku="SISTEMA",
                tipo="Erro",
                mensagem=f"Erro na limpeza: {str(e)}",
                incluir_traceback=True
            )
            
            return self._criar_resultado(
                sucesso=False,
                tempo=time.time() - inicio,
                mensagem=f"❌ Erro durante limpeza: {str(e)}\n\n{traceback.format_exc()}"
            )
    
    def _identificar_colunas_texto(self, mapa_template: Dict[str, List[int]]) -> List[int]:
        """
        Identifica colunas que contêm texto a ser limpo.
        
        Args:
            mapa_template: Mapeamento de colunas do template
            
        Returns:
            Lista de índices de colunas de texto
        """
        colunas = []
        palavras_chave = self.config.palavras_chave_texto_limpeza
        
        for nome_coluna, indices in mapa_template.items():
            if any(palavra in nome_coluna for palavra in palavras_chave):
                colunas.extend(indices)
        
        # Remove duplicatas
        return list(set(colunas))
    
    def _limpar_texto(self, texto: str) -> str:
        """
        Aplica todas as regras de limpeza ao texto.
        
        Args:
            texto: Texto original
            
        Returns:
            Texto limpo
        """
        resultado = texto
        
        # 1. Substituições (case-insensitive)
        for antigo, novo in self._termos_substituir.items():
            resultado = re.sub(re.escape(antigo), novo, resultado, flags=re.IGNORECASE)
        
        # 2. Remoções (case-insensitive)
        for termo in self._termos_remover:
            resultado = re.sub(re.escape(termo), '', resultado, flags=re.IGNORECASE)
        
        # 3. Limpeza de espaços
        # Substitui múltiplos espaços por um único
        resultado = re.sub(r' {2,}', ' ', resultado)
        
        # Remove espaços no início e fim
        resultado = resultado.strip()
        
        return resultado
