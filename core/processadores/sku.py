# ==============================================================================
# PROCESSADOR SKU
# ==============================================================================
# Processador para criação de planilhas Amazon via lista de SKUs.
#
# Este processador:
#   - Lê uma planilha de entrada com lista de SKUs
#   - Busca preços e descrições nas bases de dados
#   - Preenche o template com todos os dados
#   - Suporta Kits (K2-, K3-, K-)
# ==============================================================================

import io
import time
import openpyxl
import traceback
from typing import Dict, List, Optional, Any, Callable
from datetime import datetime

from .base import ProcessadorBase, ResultadoProcessamento
from ..config import Configuracoes
from ..utils import Utilitarios


class ProcessadorSKU(ProcessadorBase):
    """
    Processador para criação de planilhas via lista de SKUs.
    
    Este processador é especializado em:
    - Ler planilhas de entrada com SKU, Marca e EAN
    - Buscar dados completos nas bases
    - Preencher templates no formato Amazon
    - Gerar URLs de imagens automaticamente
    
    Herda de ProcessadorBase e implementa a lógica específica
    para processamento por SKU.
    """
    
    # Nomes possíveis das colunas no arquivo de entrada
    COLUNAS_SKU_INPUT = ['SKU', 'SKU-SELLER', 'ITEM', 'CÓDIGO', 'SKU SELLER']
    COLUNAS_MARCA_INPUT = ['MARCA', 'FABRICANTE', 'BRAND', 'MARCA DA EMPRESA']
    COLUNAS_EAN_INPUT = ['EAN', 'ID', 'CODIGO DE BARRAS', 'EAN DA EMPRESA']
    
    def __init__(self, config: Optional[Configuracoes] = None):
        """
        Inicializa o processador SKU.
        
        Args:
            config: Configurações do sistema (opcional)
        """
        super().__init__(config)
        self._mapa_input: Dict[str, int] = {}
    
    def _obter_nome_arquivo_saida(self) -> str:
        """Retorna o nome do arquivo de saída."""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        return f"NOGORA_PROCESSADO_{timestamp}.xlsm"
    
    def processar(self, arquivo_entrada: Any, arquivo_template: Any,
                  callback_status: Optional[Callable[[str], None]] = None,
                  callback_progresso: Optional[Callable[[float], None]] = None) -> ResultadoProcessamento:
        """
        Processa a lista de SKUs e gera o arquivo preenchido.
        
        Args:
            arquivo_entrada: Planilha com lista de SKUs
            arquivo_template: Template XLSM para preenchimento
            callback_status: Função para reportar status textual
            callback_progresso: Função para reportar progresso (0.0 a 1.0)
            
        Returns:
            Objeto ResultadoProcessamento com o resultado
        """
        inicio = time.time()
        self.limpar_logs()
        
        def status(msg: str):
            if callback_status:
                callback_status(msg)
        
        def progresso(valor: float):
            if callback_progresso:
                callback_progresso(valor)
        
        try:
            # 1. Carregar bases de dados
            status("📚 Carregando tabelas de Preços e Descrição...")
            self._carregar_bases_dados(callback_status)
            
            # 2. Abrir arquivos
            status("📂 Lendo arquivos de entrada e modelo...")
            
            # Entrada (SKUs)
            wb_entrada = openpyxl.load_workbook(arquivo_entrada, data_only=True)
            ws_entrada = wb_entrada.active
            
            # Template (Saída)
            wb_saida, ws_saida = self._abrir_template(arquivo_template, callback_status)
            
            # 3. Mapear colunas
            status("🗺️ Mapeando colunas dinamicamente...")
            
            # Mapeia entrada (linha 1)
            self._mapa_input = self.mapeador.mapear_input(ws_entrada, linha_cabecalho=1)
            
            # Mapeia template (linha 4)
            mapa_template = self._mapear_template(ws_saida, callback_status)
            
            # Verifica coluna de SKU no input
            idx_sku = self._encontrar_coluna_input(self.COLUNAS_SKU_INPUT)
            if idx_sku is None:
                raise ValueError(
                    f"Coluna de SKU não encontrada no arquivo de entrada. "
                    f"Colunas disponíveis: {list(self._mapa_input.keys())}"
                )
            
            idx_marca = self._encontrar_coluna_input(self.COLUNAS_MARCA_INPUT)
            idx_ean = self._encontrar_coluna_input(self.COLUNAS_EAN_INPUT)
            
            # 4. Processar linhas
            status("⚙️ Processando dados...")
            
            linhas = list(ws_entrada.iter_rows(min_row=2, values_only=True))
            total_linhas = len(linhas) if linhas else 1
            total_processados = 0
            
            for indice, linha in enumerate(linhas):
                # Atualiza progresso
                progresso((indice + 1) / total_linhas)
                
                # Processa a linha
                resultado_linha = self._processar_linha(
                    linha=linha,
                    idx_sku=idx_sku,
                    idx_marca=idx_marca,
                    idx_ean=idx_ean,
                    mapa_template=mapa_template
                )
                
                if resultado_linha:
                    total_processados += 1
                    self._linha_atual += 1
            
            # 5. Gerar arquivo de saída
            status("💾 Gerando arquivo de saída...")
            arquivo_saida = self._gerar_arquivo_saida()
            
            tempo_total = time.time() - inicio
            
            return self._criar_resultado(
                sucesso=True,
                arquivo=arquivo_saida,
                total_processados=total_processados,
                tempo=tempo_total,
                mensagem=f"✅ Processamento concluído! {total_processados} SKUs processados."
            )
            
        except FileNotFoundError as e:
            return self._criar_resultado(
                sucesso=False,
                tempo=time.time() - inicio,
                mensagem=f"❌ Arquivo não encontrado: {str(e)}"
            )
            
        except ValueError as e:
            return self._criar_resultado(
                sucesso=False,
                tempo=time.time() - inicio,
                mensagem=f"❌ Erro de validação: {str(e)}"
            )
            
        except Exception as e:
            # Log detalhado do erro
            self._adicionar_erro(
                sku="SISTEMA",
                tipo="Erro",
                mensagem=f"Erro fatal: {str(e)}",
                incluir_traceback=True
            )
            
            return self._criar_resultado(
                sucesso=False,
                tempo=time.time() - inicio,
                mensagem=f"❌ Erro inesperado: {str(e)}\n\n{traceback.format_exc()}"
            )
    
    def _encontrar_coluna_input(self, nomes_possiveis: List[str]) -> Optional[int]:
        """
        Encontra o índice de uma coluna na planilha de entrada.
        
        Args:
            nomes_possiveis: Lista de nomes possíveis para a coluna
            
        Returns:
            Índice da coluna ou None
        """
        return self.mapeador.encontrar_coluna_input(self._mapa_input, nomes_possiveis)
    
    def _processar_linha(self, linha: tuple, idx_sku: int, 
                        idx_marca: Optional[int], idx_ean: Optional[int],
                        mapa_template: Dict[str, List[int]]) -> bool:
        """
        Processa uma linha do arquivo de entrada.
        
        Args:
            linha: Tupla com valores da linha
            idx_sku: Índice da coluna de SKU
            idx_marca: Índice da coluna de marca (opcional)
            idx_ean: Índice da coluna de EAN (opcional)
            mapa_template: Mapeamento de colunas do template
            
        Returns:
            True se processou com sucesso
        """
        # Extrai SKU
        sku_original = linha[idx_sku] if idx_sku < len(linha) else None
        if not sku_original:
            return False
        
        sku_tratado = Utilitarios.tratar_sku(
            sku_original,
            list(self.config.mapa_prefixo_conta.keys())
        )
        
        # Extrai marca
        marca = "Genérico"
        if idx_marca is not None and idx_marca < len(linha) and linha[idx_marca]:
            marca = str(linha[idx_marca]).strip()
        
        # Extrai EAN do input
        ean_input = ""
        if idx_ean is not None and idx_ean < len(linha) and linha[idx_ean]:
            ean_input = str(linha[idx_ean]).strip()
        
        # Monta dicionário de dados
        dados = self._montar_dados_produto(sku_original, sku_tratado, marca, ean_input)
        
        # Escreve no template
        self._escrever_dados_template(dados, mapa_template)
        
        return True
    
    def _montar_dados_produto(self, sku_original: str, sku_tratado: str,
                              marca: str, ean_input: str) -> Dict[str, Any]:
        """
        Monta o dicionário completo de dados do produto.
        
        Args:
            sku_original: SKU original completo
            sku_tratado: SKU tratado (sem prefixos)
            marca: Nome da marca
            ean_input: EAN do arquivo de entrada
            
        Returns:
            Dicionário com todos os dados do produto
        """
        # Começa com valores fixos
        dados = self.config.valores_fixos_padrao.copy()
        
        # Preço
        preco = self.carregador_precos.obter_preco(sku_original)
        if preco:
            dados['preço padrão brl (vender na amazon, br)'] = preco
        else:
            self._adicionar_erro(sku_original, "Aviso", "Preço não encontrado")
        
        # Descrição
        produto = self.carregador_descricao.obter_produto(sku_tratado)
        
        # EAN (prioridade: input > descrição)
        ean_final = ean_input
        if not ean_final and produto and produto.ean:
            ean_final = produto.ean
        
        # Dados básicos
        dados.update({
            'sku': sku_original,
            'nome da marca': marca,
            'fabricante': marca,
            'id do produto': ean_final,
            'id do produto externo': ean_final
        })
        
        if produto:
            peso = Utilitarios.formatar_decimal(produto.peso)
            
            dados.update({
                'Descrição do Produto': produto.descricao,
                'número da peça do fabricante': produto.modelo,
                'número do modelo': produto.modelo,
                'número da peça': produto.modelo,
                'peso do pacote': peso,
                'peso do item': peso,
                'comprimento do pacote': Utilitarios.formatar_decimal(produto.comprimento),
                'largura do pacote': Utilitarios.formatar_decimal(produto.largura),
                'altura do pacote': Utilitarios.formatar_decimal(produto.altura),
                'comprimento': Utilitarios.formatar_decimal(produto.comprimento),
                'largura': Utilitarios.formatar_decimal(produto.largura),
                'altura': Utilitarios.formatar_decimal(produto.altura),
            })
            
            # Título
            if produto.titulo:
                dados['nome do item'] = produto.titulo
                dados['nome do produto'] = produto.titulo
                
                if len(produto.titulo) > 200:
                    self._adicionar_erro(sku_original, "Aviso", "Título > 200 caracteres")
            else:
                self._adicionar_erro(sku_original, "Erro", "Sem Título na Descrição")
            
            # Tópicos
            dados['__topicos_lista__'] = produto.topicos
            dados['__sku_tratado__'] = sku_tratado
        else:
            self._adicionar_erro(sku_original, "Erro", "Descrição não encontrada")
            dados['__topicos_lista__'] = []
            dados['__sku_tratado__'] = sku_tratado
        
        return dados
    
    def _escrever_dados_template(self, dados: Dict[str, Any],
                                  mapa_template: Dict[str, List[int]]) -> None:
        """
        Escreve os dados do produto no template.
        
        Args:
            dados: Dicionário com dados do produto
            mapa_template: Mapeamento de colunas
        """
        linha = self._linha_atual
        sku_tratado = dados.get('__sku_tratado__', '')
        topicos = dados.get('__topicos_lista__', [])
        
        # Contador para imagens extras
        contador_imagem = 2
        contador_topico = 0
        
        for nome_coluna, indices in mapa_template.items():
            # Imagem principal
            if any(k in nome_coluna for k in self.config.palavras_chave_imagem_principal):
                url = f"{self.config.url_base_imagens}/{sku_tratado}/{sku_tratado}_01.jpg"
                celula = self._planilha_saida.cell(row=linha, column=indices[0])
                self._escrever_valor_celula(celula, url)
                continue
            
            # Imagens secundárias
            eh_imagem_extra = (
                ('outro' in nome_coluna and 'imagem' in nome_coluna) or
                'other image' in nome_coluna or
                (('imagem' in nome_coluna or 'image' in nome_coluna) and
                 not any(k in nome_coluna for k in ['principal', 'main']))
            )
            
            if eh_imagem_extra:
                for idx_col in indices:
                    if contador_imagem <= 5:
                        url = f"{self.config.url_base_imagens}/{sku_tratado}/{sku_tratado}_{contador_imagem:02d}.jpg"
                        celula = self._planilha_saida.cell(row=linha, column=idx_col)
                        self._escrever_valor_celula(celula, url)
                        contador_imagem += 1
                continue
            
            # Tópicos/Bullets
            palavras_topico = ['topico', 'bullet', 'marcador', 'ponto de destaque']
            if any(k in nome_coluna for k in palavras_topico):
                for idx_col in indices:
                    if contador_topico < len(topicos):
                        celula = self._planilha_saida.cell(row=linha, column=idx_col)
                        self._escrever_valor_celula(celula, topicos[contador_topico])
                        contador_topico += 1
                continue
            
            # Outros campos
            valor = None
            
            # Busca por nome exato
            if nome_coluna in dados:
                valor = dados[nome_coluna]
            else:
                # Busca por nome normalizado
                for chave, v in dados.items():
                    if Utilitarios.normalizar_texto(chave) == nome_coluna:
                        valor = v
                        break
            
            # Escreve o valor (preservando estilo)
            if valor is not None:
                apenas_primeira = self._verificar_deve_preencher_apenas_primeira(nome_coluna)
                
                if apenas_primeira:
                    celula = self._planilha_saida.cell(row=linha, column=indices[0])
                    self._escrever_valor_celula(celula, valor)
                else:
                    for idx_col in indices:
                        celula = self._planilha_saida.cell(row=linha, column=idx_col)
                        self._escrever_valor_celula(celula, valor)
