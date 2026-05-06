# ==============================================================================
# PROCESSADOR ASIN
# ==============================================================================
# Processador para criação de planilhas Amazon via lista de ASINs.
#
# Este processador:
#   - Lê uma planilha de entrada com ASIN e SKU
#   - Busca preços e medidas nas bases de dados
#   - Preenche o template de forma simplificada (menos campos que SKU)
# ==============================================================================

import io
import time
import openpyxl
import traceback
from typing import Dict, List, Optional, Any, Callable
from datetime import datetime

from .base import ProcessadorBase, ResultadoProcessamento
from ..config import Configuracoes
from ..utils import Utilitarios


class ProcessadorASIN(ProcessadorBase):
    """
    Processador para criação de planilhas via lista de ASINs.
    
    Este processador é especializado em:
    - Ler planilhas com ASIN (coluna A) e SKU (coluna B)
    - Preencher templates com dados básicos e medidas
    - Formato simplificado comparado ao processador SKU
    
    Herda de ProcessadorBase e implementa a lógica específica
    para processamento por ASIN.
    """
    
    # Linha inicial de escrita para template ASIN (diferente do padrão)
    LINHA_INICIO_DADOS_PADRAO = 7
    
    # Nomes possíveis das colunas no arquivo de entrada
    COLUNAS_ASIN_INPUT = ['ASIN', 'CODIGO', 'CODIGO_ASIN', 'PRODUCT', 'PRODUCT CODE', 'CODE']
    COLUNAS_SKU_INPUT = ['SKU', 'SKU-SELLER', 'ITEM', 'CÓDIGO', 'SKU SELLER']
    
    # Campos que devem ser preenchidos no modo ASIN
    CAMPOS_ASIN = [
        'ASIN',
        'SKU',
        'Condição do Produto',
        'Código do canal de processamento (BR)',
        'Quantidade (BR)',
        'País de origem',
        'Baterias são necessárias?',
        'Regulamentações de produtos perigosos',
        'Unidade de comprimento do pacote',
        'Unidade de largura do pacote',
        'Unidade de altura do pacote',
        'Unidade de peso do pacote',
        'Unidade de peso do item',
        'Grupo de envio de mercadorias (BR)',
        'Peso do pacote',
        'Peso do item',
        'Comprimento do pacote',
        'Largura do pacote',
        'Altura do pacote',
        'Preço padrão BRL (Vender na Amazon, BR)'
    ]
    
    def __init__(self, config: Optional[Configuracoes] = None):
        """
        Inicializa o processador ASIN.
        
        Args:
            config: Configurações do sistema (opcional)
        """
        super().__init__(config)
        self._mapa_input: Dict[str, int] = {}
    
    def _obter_nome_arquivo_saida(self) -> str:
        """Retorna o nome do arquivo de saída."""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        return f"ListaASINS_{timestamp}.xlsm"
    
    def processar(self, arquivo_entrada: Any, arquivo_template: Any,
                  callback_status: Optional[Callable[[str], None]] = None,
                  callback_progresso: Optional[Callable[[float], None]] = None) -> ResultadoProcessamento:
        """
        Processa a lista de ASINs e gera o arquivo preenchido.
        
        Args:
            arquivo_entrada: Planilha com ASIN e SKU
            arquivo_template: Template XLSM para preenchimento
            callback_status: Função para reportar status textual
            callback_progresso: Função para reportar progresso (0.0 a 1.0)
            
        Returns:
            Objeto ResultadoProcessamento com o resultado
        """
        inicio = time.time()
        self.limpar_logs()
        
        def status(msg: str):
            if callback_status:
                callback_status(msg)
        
        def progresso(valor: float):
            if callback_progresso:
                callback_progresso(valor)
        
        try:
            # 1. Carregar bases de dados
            status("📚 Carregando tabelas de Preços e Descrição...")
            self._carregar_bases_dados(callback_status)
            
            # 2. Abrir arquivos
            status("📂 Lendo arquivos de entrada e modelo...")
            
            # Entrada (ASINs)
            wb_entrada = openpyxl.load_workbook(arquivo_entrada, data_only=True)
            ws_entrada = wb_entrada.active
            
            # Template (Saída)
            wb_saida, ws_saida = self._abrir_template(arquivo_template, callback_status)
            
            # 3. Mapear colunas do template
            status("🗺️ Mapeando colunas do template (por nome)...")
            mapa_template = self._mapear_template(ws_saida, callback_status)
            
            # 4. Mapear colunas do input de forma dinâmica
            status("🗺️ Mapeando colunas do arquivo de entrada...")
            self._mapa_input = self.mapeador.mapear_input(ws_entrada, linha_cabecalho=1)
            
            # Encontra as colunas no input
            idx_asin = self._encontrar_coluna_input(self.COLUNAS_ASIN_INPUT)
            idx_sku = self._encontrar_coluna_input(self.COLUNAS_SKU_INPUT)
            
            if idx_asin is None:
                raise ValueError(
                    f"Coluna de ASIN não encontrada no arquivo de entrada. "
                    f"Colunas disponíveis: {list(self._mapa_input.keys())}"
                )
            
            # 5. Processar linhas
            status("⚙️ Processando linhas...")
            
            linhas = list(ws_entrada.iter_rows(min_row=2, values_only=True))
            total_linhas = len(linhas) if linhas else 1
            total_processados = 0
            
            # Linha inicial específica do template ASIN
            self._linha_atual = self.LINHA_INICIO_DADOS_PADRAO
            
            for indice, linha in enumerate(linhas):
                # Atualiza progresso
                progresso((indice + 1) / total_linhas)
                
                # Extrai ASIN e SKU usando índices mapeados dinamicamente
                asin = linha[idx_asin] if idx_asin < len(linha) else None
                sku_original = linha[idx_sku] if idx_sku is not None and idx_sku < len(linha) else None
                
                if not asin:
                    continue
                
                # Processa a linha
                resultado_linha = self._processar_linha_asin(
                    asin=asin,
                    sku_original=sku_original,
                    mapa_template=mapa_template
                )
                
                if resultado_linha:
                    total_processados += 1
                    self._linha_atual += 1
            
            # 5. Gerar arquivo de saída
            status("💾 Gerando arquivo de saída...")
            arquivo_saida = self._gerar_arquivo_saida()
            
            tempo_total = time.time() - inicio
            
            return self._criar_resultado(
                sucesso=True,
                arquivo=arquivo_saida,
                total_processados=total_processados,
                tempo=tempo_total,
                mensagem=f"✅ Processamento concluído! {total_processados} ASINs processados."
            )
            
        except FileNotFoundError as e:
            return self._criar_resultado(
                sucesso=False,
                tempo=time.time() - inicio,
                mensagem=f"❌ Arquivo não encontrado: {str(e)}"
            )
            
        except Exception as e:
            self._adicionar_erro(
                sku="SISTEMA",
                tipo="Erro",
                mensagem=f"Erro fatal: {str(e)}",
                incluir_traceback=True
            )
            
            return self._criar_resultado(
                sucesso=False,
                tempo=time.time() - inicio,
                mensagem=f"❌ Erro inesperado: {str(e)}\n\n{traceback.format_exc()}"
            )
    
    def _processar_linha_asin(self, asin: str, sku_original: Optional[str],
                               mapa_template: Dict[str, List[int]]) -> bool:
        """
        Processa uma linha do arquivo de entrada ASIN.
        
        Args:
            asin: Código ASIN
            sku_original: SKU original
            mapa_template: Mapeamento de colunas do template
            
        Returns:
            True se processou com sucesso
        """
        linha = self._linha_atual
        
        # Monta dados (obter_produto faz seu próprio tratar_sku internamente)
        dados = self._montar_dados_asin(asin, sku_original)
        
        # Pré-normaliza as chaves dos dados para matching direto com o template
        dados_normalizados = {}
        for chave, valor in dados.items():
            chave_norm = Utilitarios.normalizar_texto(chave)
            dados_normalizados[chave_norm] = valor
        
        # Escreve no template
        for nome_coluna_template, indices in mapa_template.items():
            valor = dados_normalizados.get(nome_coluna_template)
            
            if valor is None:
                continue
            
            # Verifica se deve preencher apenas a primeira
            apenas_primeira = self._verificar_deve_preencher_apenas_primeira(nome_coluna_template)
            
            # Escreve preservando estilos
            if apenas_primeira:
                celula = self._planilha_saida.cell(row=linha, column=indices[0])
                self._escrever_valor_celula(celula, valor)
            else:
                for idx_col in indices:
                    celula = self._planilha_saida.cell(row=linha, column=idx_col)
                    self._escrever_valor_celula(celula, valor)
        
        return True
    
    def _montar_dados_asin(self, asin: str, 
                           sku_original: Optional[str]) -> Dict[str, Any]:
        """
        Monta o dicionário de dados para processamento ASIN.
        
        Args:
            asin: Código ASIN
            sku_original: SKU original completo (com prefixo de conta)
            
        Returns:
            Dicionário com dados do produto
        """
        dados = {}
        
        # Identificação
        dados['ASIN'] = asin
        dados['SKU'] = sku_original or ""
        
        # Valores fixos (apenas os necessários para ASIN)
        campos_fixos = [
            'Condição do Produto',
            'Código do canal de processamento (BR)',
            'Quantidade (BR)',
            'País de origem',
            'Baterias são necessárias?',
            'Regulamentações de produtos perigosos',
            'Unidade de comprimento do pacote',
            'Unidade de largura do pacote',
            'Unidade de altura do pacote',
            'Unidade de peso do pacote',
            'Unidade de peso do item',
            'Grupo de envio de mercadorias (BR)'
        ]
        
        for campo in campos_fixos:
            if campo in self.config.valores_fixos_padrao:
                dados[campo] = self.config.valores_fixos_padrao[campo]
            else:
                campo_norm = Utilitarios.normalizar_texto(campo)
                for chave, valor in self.config.valores_fixos_padrao.items():
                    if Utilitarios.normalizar_texto(chave) == campo_norm:
                        dados[campo] = valor
                        break
        
        # Preço (obter_preco já faz tratar_sku internamente)
        if sku_original:
            preco = self.carregador_precos.obter_preco(sku_original)
            if preco:
                dados['Preço padrão BRL (Vender na Amazon, BR)'] = preco
            else:
                self._adicionar_erro(sku_original or asin, "Aviso", "Preço não encontrado")
        
        # Medidas e Pesos (obter_produto já faz tratar_sku e cálculo de kit internamente)
        if sku_original:
            produto = self.carregador_descricao.obter_produto(sku_original)
            
            if produto:
                peso = Utilitarios.formatar_decimal(produto.peso)
                comp = Utilitarios.formatar_decimal(produto.comprimento)
                larg = Utilitarios.formatar_decimal(produto.largura)
                alt = Utilitarios.formatar_decimal(produto.altura)
                
                dados['Peso do pacote'] = peso
                dados['Peso do item'] = peso
                dados['Comprimento do pacote'] = comp
                dados['Largura do pacote'] = larg
                dados['Altura do pacote'] = alt
                
                self._adicionar_erro(
                    sku_original or asin,
                    "Info",
                    f"Medidas: peso={peso} comp={comp} larg={larg} alt={alt}"
                )
            else:
                self._adicionar_erro(
                    sku_original or asin, 
                    "Aviso", 
                    f"Descricao/medidas nao encontradas na base"
                )
        
        return dados
    
    def _encontrar_coluna_input(self, nomes_possiveis: List[str]) -> Optional[int]:
        """
        Encontra o índice de uma coluna na planilha de entrada.
        
        Args:
            nomes_possiveis: Lista de nomes possíveis para a coluna
            
        Returns:
            Índice da coluna ou None
        """
        return self.mapeador.encontrar_coluna_input(self._mapa_input, nomes_possiveis)
